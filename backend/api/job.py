from flask import Blueprint, request, jsonify, send_file, make_response, current_app, url_for
from backend.models.driver_remark import DriverRemark
from backend.services.job_service import JobService, ServiceError
from backend.services.bill_service import BillService, ServiceError as BillServiceError
from backend.schemas.job_schema import JobSchema
from backend.utils.validation import validate_job_row, get_validation_lookups, validate_excel_data
import requests
import logging
import tempfile
import os
import json
import pandas as pd
from datetime import datetime, timedelta
import io
from io import BytesIO
from flask_security.decorators import roles_accepted, auth_required
from flask_security import current_user
from werkzeug.utils import secure_filename
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl import load_workbook
import re
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
from sqlalchemy import func, or_, and_
from flask import abort

from backend.models.job import Job, JobStatus
from backend.models.customer import Customer
from backend.models.driver import Driver
from backend.models.service import Service
from backend.models.vehicle import Vehicle
from backend.models.invoice import Invoice
from backend.models.job_audit import JobAudit
from backend.models.job_photo import JobPhoto
from backend.models.user import User
from backend.models.contractor import Contractor
from backend.models.vehicle_type import VehicleType
from backend.extensions import db
from decimal import Decimal

limiter = Limiter(key_func=get_remote_address)

job_bp = Blueprint('job', __name__)
schema = JobSchema(session=db.session)
schema_many = JobSchema(many=True, session=db.session)

def sanitize_filter_value(value):
    """
    Sanitize filter values to prevent SQL injection.
    Only allows alphanumeric characters, spaces, hyphens, underscores, and common punctuation.
    """
    if not value:
        return None
    
    # Remove any SQL injection attempts
    value = str(value).strip()
    
    # Check for SQL injection patterns first
    sql_injection_patterns = [
        r'[\'";]',  # Single quotes, double quotes, semicolons
        r'--',      # SQL comments
        r'/\*',     # SQL block comments
        r'\*/',     # SQL block comments
        r'xp_cmdshell',  # SQL Server command shell
        r'WAITFOR',      # SQL Server wait
        r'SHUTDOWN',     # SQL Server shutdown
        r'DROP\s+TABLE', # DROP TABLE
        r'DELETE\s+FROM', # DELETE FROM
        r'INSERT\s+INTO', # INSERT INTO
        r'UPDATE\s+SET',  # UPDATE SET
        r'UNION\s+SELECT', # UNION SELECT
        r'EXEC\s+',       # EXEC
        r'EXECUTE\s+',    # EXECUTE
    ]
    
    for pattern in sql_injection_patterns:
        if re.search(pattern, value, re.IGNORECASE):
            logging.warning(f"SQL injection pattern detected: {pattern} in value: {value}")
            return None
    
    # Only allow safe characters for search
    # Allow alphanumeric, spaces, hyphens, underscores, dots, and limited punctuation
    safe_pattern = re.compile(r'^[a-zA-Z0-9\s\-_.,:!?@#$%&*()+=<>[\]{}|\\/"`~]+$')
    
    if not safe_pattern.match(value):
        logging.warning(f"Potentially unsafe filter value detected: {value}")
        return None
    
    return value

def apply_safe_filter(query, column, value):
    """
    Apply a safe filter using SQLAlchemy's parameterized queries.
    """
    sanitized_value = sanitize_filter_value(value)
    if sanitized_value is None:
        return query
    
    # Use SQLAlchemy's parameterized queries to prevent SQL injection
    # Use string concatenation with sanitized value instead of f-string
    pattern = '%' + sanitized_value + '%'
    return query.filter(column.ilike(pattern))

# --- add this helper near the top of the file (below imports) ---
def scoped_jobs_query(q):
    """
    Restrict the Job query based on the current_user role.
    Admin/Manager/Accountant: see all (no filter)
    Driver: only jobs for driver_id
    Customer: only jobs for customer_id
    """
    # Admins, managers, accountants see all
    if current_user.has_role('admin') or current_user.has_role('manager') or current_user.has_role('accountant'):
        return q

    # Driver scope
    if current_user.has_role('driver'):
        driver_id = getattr(current_user, 'driver_id', None)
        if not driver_id:
            abort(403, description='Driver profile missing')
        return q.filter(Job.driver_id == driver_id)

    # Customer scope
    if current_user.has_role('customer'):
        customer_id = getattr(current_user, 'customer_id', None)
        if not customer_id:
            abort(403, description='Customer profile missing')
        return q.filter(Job.customer_id == customer_id)

    abort(403, description='Access forbidden')

# --- end helper ---

@job_bp.route('/jobs', methods=['GET'])
@auth_required()
def list_jobs():
    try:
        # use the same scope helper to ensure the same RBAC rules
        query = scoped_jobs_query(Job.query.filter(Job.is_deleted.is_(False)))
        jobs = query.all()
        return jsonify(schema_many.dump(jobs)), 200

    except ServiceError as se:
        return jsonify({'error': se.message}), 400
    except Exception as e:
        logging.error(f"Unhandled error in list_jobs: {e}", exc_info=True)
        return jsonify({'error': 'An unexpected error occurred. Please try again later.'}), 500

@job_bp.route('/jobs/<int:job_id>', methods=['GET'])
@auth_required()
def get_job(job_id):
    try:
        # enforce scope even for single job fetch
        query = scoped_jobs_query(Job.query.filter(Job.is_deleted.is_(False)))
        job = query.filter(Job.id == job_id).first()
        if not job:
            return jsonify({'error': 'Job not found'}), 404
        return jsonify(schema.dump(job)), 200

    except ServiceError as se:
        return jsonify({'error': se.message}), 400
    except Exception as e:
        logging.error(f"Unhandled error in get_job: {e}", exc_info=True)
        return jsonify({'error': 'An unexpected error occurred. Please try again later.'}), 500



@job_bp.route('/jobs', methods=['POST'])
@roles_accepted('admin', 'manager', 'accountant', 'customer')
def create_job():
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
            
        errors = schema.validate(data)
        if errors:
            return jsonify(errors), 400
            
        # Check for driver scheduling conflict
        driver_id = data.get('driver_id')
        pickup_date = data.get('pickup_date')
        pickup_time = data.get('pickup_time')
        time_buffer_minutes = data.get('time_buffer_minutes', 60)  # Default to 60 minutes
        
        conflict_job = None
        if driver_id and pickup_date and pickup_time:
            # Use transactional isolation to prevent race conditions
            # Lock the driver row to prevent concurrent scheduling
            if driver_id:
                # Use SQLAlchemy's with_for_update() method which handles database differences
                # This provides proper row-level locking for PostgreSQL/MySQL and emulates 
                # locking for SQLite via transaction serialization
                from sqlalchemy.orm import selectinload
                driver = db.session.query(Driver).filter(Driver.id == driver_id).with_for_update().one()
            
            conflict_job = JobService.check_driver_conflict(driver_id, pickup_date, pickup_time, None, time_buffer_minutes)
            # Completely override scheduling conflict validation
            # Allow all authenticated users to create jobs even with scheduling conflicts
            if conflict_job:
                logging.warning(f"Authenticated user {current_user.email} is creating a job that conflicts with job #{conflict_job.id} - OVERRIDE ENABLED")
        
        job = JobService.create(data)
        
        # Create audit record for job creation
        try:
            from datetime import datetime
            audit_record = JobAudit(
                job_id=job.id,
                changed_by=current_user.id if current_user.is_authenticated else None,
                old_status="new",
                new_status=job.status,
                reason="Job created",
                additional_data=data,  # Store all the fields that were used to create the job
                changed_at=datetime.now()
            )
            db.session.add(audit_record)
            db.session.commit()
        except Exception as audit_error:
            logging.error(f"Error creating audit record for job creation: {audit_error}", exc_info=True)
            # Don't fail the job creation if audit logging fails
            db.session.rollback()
        # Prepare response
        job_data = schema.dump(job)
       
        # Include conflict warning for admin users
        if conflict_job and current_user.has_role('admin'):
            # Create a new response object with the job data and warning
            response_data = {
                'job': job_data,
                'warning': {
                    'type': 'scheduling_conflict',
                    'message': f'Created despite conflict with Job #{conflict_job.id} ({conflict_job.status}) at {conflict_job.pickup_date} {conflict_job.pickup_time}',
                    'details': {
                        'conflict_job_id': conflict_job.id,
                        'status': conflict_job.status,
                        'pickup_date': conflict_job.pickup_date,
                        'pickup_time': conflict_job.pickup_time
                    }
                }
            }
            return jsonify(response_data), 201
        else:
            return jsonify(job_data), 201
    except ServiceError as se:
        db.session.rollback()
        return jsonify({'message': se.message}), 400
    except Exception as e:
        db.session.rollback()
        logging.error(f"Unhandled error in create_job: {e}", exc_info=True)
        return jsonify({'error': 'An unexpected error occurred. Please try again later.'}), 500

@job_bp.route('/jobs/<int:job_id>', methods=['PUT'])
@roles_accepted('admin', 'manager', 'accountant', 'customer')
def update_job(job_id):
    try:
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
            
        # Filter out read-only/dump_only fields before validation
        # These fields are computed or read-only and should not be updated
        read_only_fields = {
            'customer_name', 'customer_email', 'customer_mobile', 'customer_reference',
            'vehicle_type', 'vehicle_number', 'driver_contact',
            'payment_mode', 'message', 'remarks', 'has_additional_stop', 'additional_stops',
            'base_discount_percent', 'customer_discount_percent', 'additional_discount_percent',
            'invoice_number', 'type_of_service', 'reference', 'locations', 'has_request',
            'duration_minutes', 'duration_str', 'customer', 'driver', 'vehicle', 'service',
            'invoice', 'sub_customer', 'contractor'
        }
        
        # Remove read-only fields from data
        filtered_data = {k: v for k, v in data.items() if k not in read_only_fields}
            
        errors = schema.validate(filtered_data, partial=True)
        if errors:
            return jsonify(errors), 400
            
        # Get the current job before updating to capture old status
        job_before_update = Job.query.get(job_id)
        if not job_before_update:
            return jsonify({'error': 'Job not found'}), 404
            
        old_status = job_before_update.status
        status = (job_before_update.status or "").lower()

        if current_user.has_role("customer"):
            # Completely locked statuses
            locked_statuses = {"jc", "sd", "canceled"}
            # Limited editable statuses (remarks only)
            limited_edit_statuses = {"confirmed", "otw", "ots", "pob"}

            # Full lock-out
            if status in locked_statuses:
                return jsonify({
                    "error": "Permission denied",
                    "message": f"Customers cannot edit jobs in '{status}' status."
                }), 403
            # Partial permission: remarks only
            if status in limited_edit_statuses:
                allowed_fields = {"remarks","customer_remark"}
                def normalize(val):
                    if val in (None, "", [], {}, 0, 0.0, Decimal("0")):
                        return 0
                    if isinstance(val, (int, float, Decimal)):
                        return round(float(val), 2)
                    if isinstance(val, (list, dict)):
                        return json.dumps(val, sort_keys=True)
                    return str(val)
                changed_fields = { k for k, v in data.items() if normalize(getattr(job_before_update, k, None)) != normalize(v)}
                disallowed_fields = [f for f in changed_fields if f not in allowed_fields]

                if disallowed_fields:
                    return jsonify({
                        "error": "Permission denied",
                        "message": f"Customers can only edit remarks in '{status}' status.",
                        "disallowed_fields": disallowed_fields
                    }), 403
        # Check for driver scheduling conflict
        driver_id = filtered_data.get('driver_id')
        pickup_date = filtered_data.get('pickup_date')
        pickup_time = filtered_data.get('pickup_time')
        time_buffer_minutes = filtered_data.get('time_buffer_minutes', 60)  # Default to 60 minutes
        
        conflict_job = None
        if driver_id and pickup_date and pickup_time:
            # Use transactional isolation to prevent race conditions
            # Lock the driver row to prevent concurrent scheduling
            if driver_id:
                # Use SQLAlchemy's with_for_update() method which handles database differences
                # This provides proper row-level locking for PostgreSQL/MySQL and emulates 
                # locking for SQLite via transaction serialization
                from sqlalchemy.orm import selectinload
                driver = db.session.query(Driver).filter(Driver.id == driver_id).with_for_update().one()
            
            conflict_job = JobService.check_driver_conflict(driver_id, pickup_date, pickup_time, job_id, time_buffer_minutes)
            # Completely override scheduling conflict validation
            # Allow all authenticated users to update jobs even with scheduling conflicts
            if conflict_job:
                logging.warning(f"Authenticated user {current_user.email} is updating a job that conflicts with job #{conflict_job.id} - OVERRIDE ENABLED")
        
        job = JobService.update(job_id, filtered_data)
        if not job:
            return jsonify({'error': 'Job not found'}), 404
        
            
        # Audit logging is now handled in JobService.update to prevent duplicate audit records.
            
        # Prepare response
        job_data = schema.dump(job)
        
        
        # Include conflict warning for admin users
        if conflict_job and current_user.has_role('admin'):
            # Create a new response object with the job data and warning
            response_data = {
                'job': job_data,
                'warning': {
                    'type': 'scheduling_conflict',
                    'message': f'Updated despite conflict with Job #{conflict_job.id} ({conflict_job.status}) at {conflict_job.pickup_date} {conflict_job.pickup_time}',
                    'details': {
                        'conflict_job_id': conflict_job.id,
                        'status': conflict_job.status,
                        'pickup_date': conflict_job.pickup_date,
                        'pickup_time': conflict_job.pickup_time
                    }
                }
            }
            return jsonify(response_data), 200
        else:
            return jsonify(job_data), 200
    except ServiceError as se:
        db.session.rollback()
        return jsonify({'message': se.message}), 400
    except Exception as e:
        db.session.rollback()
        logging.error(f"Unhandled error in update_job: {e}", exc_info=True)
        return jsonify({'error': 'An unexpected error occurred. Please try again later.'}), 500

@job_bp.route('/jobs/<int:job_id>', methods=['DELETE'])
@roles_accepted('admin', 'manager', 'accountant')
def delete_job(job_id):
    """
    Delete a job (soft delete implementation).
    
    This endpoint implements soft delete by directly setting the is_deleted flag 
    to True with proper transaction boundaries and idempotency checks. This ensures 
    data integrity and allows for audit trails and potential recovery.
    
    Args:
        job_id (int): The ID of the job to delete
        
    Returns:
        JSON response with success message or error
    """
    try:
        # Use with_for_update for pessimistic locking to prevent race conditions
        job = Job.query.filter_by(id=job_id).with_for_update().first()
        if not job:
            return jsonify({'error': 'Job not found'}), 404

        # Idempotency check - if job is already deleted, return success
        if job.is_deleted:
            return jsonify({'message': 'Job already marked as deleted', 'already_deleted': True}), 200

        # Check if job can be deleted (e.g., not in certain statuses)
        # Note: This is an example - adjust business rules as needed
        if job.status in [JobStatus.POB.value]:  # Example: can't delete in-progress jobs
            return jsonify({'error': 'Cannot delete job in current status'}), 400

        # Mark job as deleted
        job.is_deleted = True
        db.session.commit()
        
        return jsonify({'message': 'Job deleted', 'already_deleted': False}), 200
    except Exception as e:
        db.session.rollback()
        logging.error(f"Unhandled error in delete_job: {e}", exc_info=True)
        return jsonify({'error': 'An unexpected error occurred. Please try again later.'}), 500


# @job_bp.route('/jobs/calculate_price', methods=['POST'])
# def calculate_job_price():
#     try:
#         data = request.get_json()
#         required_fields = ['customer_id', 'vehicle_type', 'service_type']
#         for field in required_fields:
#             if field not in data:
#                 return jsonify({'error': f'Missing required field: {field}'}), 400
#         result = JobService.calculate_price(
#             customer_id=data['customer_id'],
#             vehicle_type=data['vehicle_type'],
#             service_type=data['service_type'],
#             year=data.get('year'),
#             extra_services=data.get('extra_services'),
#             additional_stop_count=data.get('additional_stop_count', 0),
#             midnight=data.get('midnight', False)
#         )
#         if isinstance(result, dict) and result.get('error'):
#             return jsonify(result), 400
#         return jsonify(result), 200
#     except ServiceError as se:
#         return jsonify({'error': se.message}), 400
#     except Exception as e:
#         logging.error(f"Unhandled error in calculate_job_price: {e}", exc_info=True)
#         return jsonify({'error': 'An unexpected error occurred. Please try again later.'}), 500


@job_bp.route('/jobs/calculate_price', methods=['POST'])
def calculate_job_price():
    try:
        data = request.get_json()
        required_fields = ['customer_id', 'vehicle_type', 'service_type']
        for field in required_fields:
            if field not in data:
                return jsonify({'error': f'Missing required field: {field}'}), 400

        # Pass full data dict
        result = JobService.calculate_price(data)

        if isinstance(result, dict) and result.get('error'):
            return jsonify(result), 400
        return jsonify(result), 200

    except ServiceError as se:
        return jsonify({'error': se.message}), 400
    except Exception as e:
        logging.error(f"Unhandled error in calculate_job_price: {e}", exc_info=True)
        return jsonify({'error': 'An unexpected error occurred. Please try again later.'}), 500



@job_bp.route('/jobs/<int:job_id>/calculate_driver_commission', methods=['POST'])
def calculate_job_driver_commission(job_id):
    try:
        result = JobService.calculate_driver_commission(job_id)
        if isinstance(result, dict) and result.get('error'):
            return jsonify(result), 400
        return jsonify(result), 200
    except ServiceError as se:
        return jsonify({'error': se.message}), 400
    except Exception as e:
        logging.error(f"Unhandled error in calculate_job_driver_commission: {e}", exc_info=True)
        return jsonify({'error': 'An unexpected error occurred. Please try again later.'}), 500

@job_bp.route('/jobs/<int:job_id>/penalty', methods=['POST'])
def set_job_penalty(job_id):
    try:
        data = request.get_json()
        penalty = data.get('penalty')
        if penalty is None:
            return jsonify({'error': 'penalty is required'}), 400
        result = JobService.set_penalty(job_id, penalty)
        if isinstance(result, dict) and result.get('error'):
            return jsonify(result), 400
        return jsonify({'message': 'Penalty updated', 'job_id': job_id, 'penalty': penalty}), 200
    except ServiceError as se:
        return jsonify({'message': se.message}), 400
    except Exception as e:
        logging.error(f"Unhandled error in set_job_penalty: {e}", exc_info=True)
        return jsonify({'error': 'An unexpected error occurred. Please try again later.'}), 500


@job_bp.route('/jobs/<int:job_id>/soft-delete', methods=['PUT'])
@roles_accepted('admin', 'manager', 'accountant')
def soft_delete_job(job_id):
    """
    Soft delete a job with proper transaction boundaries and idempotency checks.
    
    This endpoint marks a job as deleted by setting the is_deleted flag to True.
    It includes proper transaction handling, idempotency checks, and business
    rule validation to prevent race conditions and ensure data integrity.
    
    Args:
        job_id (int): The ID of the job to soft delete
        
    Returns:
        JSON response with success message or error
    """
    try:
        # Use with_for_update for pessimistic locking to prevent race conditions
        job = Job.query.filter_by(id=job_id).with_for_update().first()
        if not job:
            return jsonify({'error': 'Job not found'}), 404

        # Idempotency check - if job is already deleted, return success
        if job.is_deleted:
            return jsonify({'message': 'Job already marked as deleted', 'already_deleted': True}), 200

        # Check if job can be deleted (e.g., not in certain statuses)
        # Note: This is an example - adjust business rules as needed
        if job.status in [JobStatus.POB.value]:  # Example: can't delete in-progress jobs
            return jsonify({'error': 'Cannot delete job in current status'}), 400

        # Mark job as deleted
        job.is_deleted = True
        db.session.commit()
        
        return jsonify({'message': 'Job marked as deleted', 'already_deleted': False}), 200
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error in soft_delete_job: {e}", exc_info=True)
        return jsonify({'error': 'An unexpected error occurred. Please try again later.'}), 500

@job_bp.route('/jobs/check-driver-conflict', methods=['POST'])
@auth_required()
def check_driver_conflict():
    try:
        data = request.get_json()
        driver_id = data.get('driver_id')
        pickup_date = data.get('pickup_date')
        pickup_time = data.get('pickup_time')
        job_id = data.get('job_id')  # Optional, for updates
        time_buffer_minutes = data.get('time_buffer_minutes', 60)  # Optional, default to 60 minutes
        
        if not driver_id or not pickup_date or not pickup_time:
            return jsonify({'error': 'Missing required parameters'}), 400
            
        conflict_job = JobService.check_driver_conflict(driver_id, pickup_date, pickup_time, job_id, time_buffer_minutes)
        
        if conflict_job:
            response = {
                'conflict': True,
                'message': f'Driver already has a job scheduled at {conflict_job.pickup_date} {conflict_job.pickup_time} (Job #{conflict_job.id} - {conflict_job.status}). Please select a different date or time.',
                'conflict_details': {
                    'conflict_job_id': conflict_job.id,
                    'pickup_date': conflict_job.pickup_date,
                    'pickup_time': conflict_job.pickup_time,
                    'status': conflict_job.status
                }
            }
            # Only include admin warning if user is admin
            if current_user.has_role('admin'):
                response['warning'] = 'As an administrator, you can override this conflict.'
            return jsonify(response), 200
        else:
            return jsonify({'conflict': False, 'message': 'No scheduling conflicts detected.'}), 200
            
    except Exception as e:
        logging.error(f"Unhandled error in check_driver_conflict: {e}", exc_info=True)
        return jsonify({'error': 'An unexpected error occurred. Please try again later.'}), 500

@job_bp.route('/jobs/table', methods=['GET'])
@auth_required()
def jobs_table():
    # Get jobs with filters
    try:
        # Parse filters and pagination
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('pageSize', 50))
        # Limit page size to maximum 100 results as per acceptance criteria
        page_size = min(page_size, 100)
        filters = {}
        search_term = None
        for key in request.args:
            if key not in ['page', 'pageSize']:
                value = request.args.get(key)
                if key == 'search':
                    search_term = value
                else:
                    filters[key] = value
        
        # Query with filters
        query = Job.query.filter(Job.is_deleted.is_(False))
        query = scoped_jobs_query(query)

        joined_customer = False
        joined_driver = False
        joined_vehicle = False
        joined_contractor = False
        joined_vehicle_type = False
        
        # Handle computed field filters by joining with related tables
        for key, value in filters.items():
            if value:
                if key == 'customer_name':
                    # Filter by customer name - use parameterized query with join
                    if not joined_customer:
                        query = query.join(Customer)
                        joined_customer = True
                    sanitized_value = sanitize_filter_value(value)
                    if sanitized_value:
                        pattern = '%' + sanitized_value + '%'
                        query = query.filter(Customer.name.ilike(pattern))
                elif key == 'service_type':
                    # Filter by service type - use parameterized query
                    query = apply_safe_filter(query, Job.service_type, value)
                elif key == 'pickup_location':
                    # Filter by pickup location - use parameterized query
                    query = apply_safe_filter(query, Job.pickup_location, value)
                elif key == 'dropoff_location':
                    # Filter by dropoff location - use parameterized query
                    query = apply_safe_filter(query, Job.dropoff_location, value)
                elif key == 'pickup_date':
                    # Filter by pickup date - use parameterized query
                    query = apply_safe_filter(query, Job.pickup_date, value)
                elif key == 'pickup_time':
                    # Filter by pickup time - use parameterized query
                    query = apply_safe_filter(query, Job.pickup_time, value)
                elif key == 'passenger_name':
                    # Filter by passenger name - use parameterized query
                    query = apply_safe_filter(query, Job.passenger_name, value)
                elif key == 'status':
                    # Filter by status - use exact match for enum values
                    query = query.filter(Job.status == value)
                elif key == 'vehicle_number':
                    # Filter by vehicle number - join Vehicle table if not already joined
                    if not joined_vehicle:
                        query = query.outerjoin(Vehicle, Job.vehicle_id == Vehicle.id)
                        joined_vehicle = True
                    query = apply_safe_filter(query, Vehicle.number, value)
                elif key == 'passenger_email':
                    # Filter by passenger email - use parameterized query
                    query = apply_safe_filter(query, Job.passenger_email, value)
                elif key == 'customer_email':
                    # Filter by customer email - join Customer table if not already joined
                    if not joined_customer:
                        query = query.join(Customer)
                        joined_customer = True
                    query = apply_safe_filter(query, Customer.email, value)
                elif hasattr(Job, key):
                    # For other direct Job model fields - use exact match
                    query = query.filter(getattr(Job, key) == value)
        
        # Apply comprehensive search if search term is provided
        if search_term:
            search_term_sanitized = sanitize_filter_value(search_term.strip() if search_term else '')
            if search_term_sanitized:
                # Convert search term to different date formats if needed
                converted_term = search_term_sanitized
                try:
                    from datetime import datetime
                    date_formats = ['%d/%m/%Y', '%d-%m-%Y']
                    for fmt in date_formats:
                        try:
                            search_date = datetime.strptime(search_term_sanitized, fmt)
                            converted_term = search_date.strftime('%Y-%m-%d')
                            break
                        except ValueError:
                            continue
                except Exception:
                    pass
                
                # Create a search pattern for LIKE queries
                search_pattern = f'%{converted_term}%'
                
                # For numeric-only search terms, match exact job ID and partial booking_ref (as per acceptance criteria)
                if search_term_sanitized.isdigit():
                    exact_job_id = int(search_term_sanitized)
                    search_pattern = f'%{search_term_sanitized}%'
                    query = query.filter(
                        or_(
                            Job.id == exact_job_id,
                            Job.booking_ref.ilike(search_pattern)
                        )
                    )
                else:
                    # For non-numeric searches, join tables as needed for related fields
                    if not joined_customer:
                        query = query.join(Customer)
                        joined_customer = True
                    if not joined_driver:
                        query = query.outerjoin(Driver, Job.driver_id == Driver.id)
                        joined_driver = True
                    if not joined_vehicle:
                        query = query.outerjoin(Vehicle, Job.vehicle_id == Vehicle.id)
                        joined_vehicle = True
                    if not joined_contractor:
                        query = query.outerjoin(Contractor, Job.contractor_id == Contractor.id)
                        joined_contractor = True
                    if not joined_vehicle_type:
                        query = query.outerjoin(VehicleType, Job.vehicle_type_id == VehicleType.id)
                        joined_vehicle_type = True
                    
                    # Apply search across multiple fields using OR condition
                    search_conditions = [
                        Job.id.cast(db.String).ilike(search_pattern),  # job_id
                        Job.passenger_name.ilike(search_pattern),  # passenger_name
                        func.coalesce(Job.passenger_email, '').ilike(search_pattern),  # passenger_email
                        func.coalesce(Job.passenger_mobile, '').ilike(search_pattern),  # passenger_mobile
                        Job.booking_ref.ilike(search_pattern),  # booking_ref
                        Job.pickup_location.ilike(search_pattern),  # pickup_location
                        Job.dropoff_location.ilike(search_pattern),  # dropoff_location
                        func.coalesce(Job.service_type, '').ilike(search_pattern),  # service_type
                        Job.pickup_date.ilike(search_pattern),  # pickup_date
                        func.coalesce(Job.pickup_time, '').ilike(search_pattern),  # pickup_time
                        Customer.name.ilike(search_pattern),  # customer_name
                        func.coalesce(Customer.email, '').ilike(search_pattern),  # customer_email
                        func.coalesce(Job.sub_customer_name, '').ilike(search_pattern),  # sub_customer_name
                        func.coalesce(Driver.name, '').ilike(search_pattern),  # driver_name
                        func.coalesce(Vehicle.number, '').ilike(search_pattern),  # vehicle_name
                        func.coalesce(Contractor.name, '').ilike(search_pattern),  # contractor_name
                        func.coalesce(VehicleType.name, '').ilike(search_pattern)  # vehicle_type_name
                    ]
                    
                    # For non-numeric searches, apply search across multiple fields
                    query = query.filter(or_(*search_conditions))
        
        total = query.count()
        
        # Use optimized loading for table view - load only essential relationships
        optimized_query = query.options(
            db.joinedload(Job.customer),
            db.joinedload(Job.driver),
            db.joinedload(Job.vehicle),
            db.joinedload(Job.vehicle_type)
        )
        
        jobs = optimized_query.order_by(Job.pickup_date.desc(), Job.pickup_time.desc()).offset((page-1)*page_size).limit(page_size).all()
        
        job_items = schema_many.dump(jobs)

        # Inject driver_remark if driver_id param exists
        driver_id_filter = request.args.get("driver_id", type=int)
        job_id_filter = request.args.get("id", type=int)

        if driver_id_filter and job_id_filter:
            remarks = DriverRemark.query.filter_by(job_id=job_id_filter, driver_id=driver_id_filter) \
                .order_by(DriverRemark.created_at.asc()).all()
            if remarks:
                for item in job_items:
                    if isinstance(item, dict) and item.get("id") == job_id_filter:
                        item["driver_remarks"] = [
                            {
                                "id": r.id,
                                "remark": r.remark,
                                "created_at": r.created_at.isoformat()
                            }
                            for r in remarks
                ]
        # Determine if results were limited
        # isLimited should indicate if the results were artificially capped
        is_limited = len(job_items) == page_size and page_size >= 100
        
        return jsonify({
            'items': job_items,
            'total': total,
            'page': page,
            'pageSize': page_size,
            'isLimited': is_limited,
            'search': search_term or ''
        }), 200
    except Exception as e:
        logging.error(f"Unhandled error in jobs_table: {e}", exc_info=True)
        return jsonify({'error': 'An unexpected error occurred. Please try again later.'}), 500

@job_bp.route('/jobs/template', methods=['GET'])
@auth_required()
def download_job_template():
    """Download Excel template for bulk job uploads"""
    try:
        # Check if user is a Customer role
        is_customer_user = current_user.has_role('customer')

        # Get sample data for dropdowns - filter based on user role
        if is_customer_user:
            # For Customer users: Only show their own customer
            customer_id = getattr(current_user, 'customer_id', None)
            if not customer_id:
                logging.error(f"Customer user {current_user.email} has no customer_id configured")
                return jsonify({
                    'error': 'Your account is not properly configured. Please contact your administrator to set up your customer association.'
                }), 400

            customers = Customer.query.filter_by(id=customer_id, status='Active').all()
            if not customers:
                logging.error(f"Customer user {current_user.email} has invalid customer_id: {customer_id}")
                return jsonify({
                    'error': 'Your associated customer account is not found or inactive. Please contact your administrator.'
                }), 400

            # For Customer users: Empty vehicles and drivers (no values in dropdown)
            vehicles = []
            drivers = []

            # For Customer users: Only show AG (Internal) contractor
            # Try to find contractor with name 'AG' or 'AG (Internal)'
            contractors = Contractor.query.filter(
                Contractor.status == 'Active',
                Contractor.name.in_(['AG', 'AG (Internal)'])
            ).all()
            # If not found, try partial match for any contractor containing 'AG'
            if not contractors:
                contractors = Contractor.query.filter(
                    Contractor.status == 'Active',
                    Contractor.name.like('%AG%')
                ).all()
            # Handle case where no AG Internal contractor exists
            if not contractors:
                logging.warning("No AG Internal contractor found for customer template - samples may be invalid")
                contractors = []
            logging.info(f"Customer user - Found {len(contractors)} AG contractors: {[c.name for c in contractors]}")
        else:
            # For Admin/Manager users: Show all active data
            customers = Customer.query.filter_by(status='Active').all()
            vehicles = Vehicle.query.filter_by(status='Active').all()
            drivers = Driver.query.filter_by(status='Active').all()
            contractors = Contractor.query.filter_by(status='Active').all()

        # Services and Vehicle Types are same for all users
        services = Service.query.filter_by(status='Active').all()
        vehicle_types = VehicleType.query.filter_by(status=True, is_deleted=False).all()
        
        # Create sample data using actual database values
        sample_data = []
        today = datetime.now()

        # Pre-check and extract safe values for optional fields
        contractor_value = contractors[0].name if contractors else ''
        vehicle_type_value = vehicle_types[0].name if vehicle_types else ''
        contractor_value_2 = contractors[1].name if len(contractors) > 1 else contractor_value
        vehicle_type_value_2 = vehicle_types[1].name if len(vehicle_types) > 1 else vehicle_type_value

        # Create valid sample data based on user role
        if is_customer_user:
            # For Customer users: Create sample data WITHOUT Vehicle/Driver columns
            if customers and services and len(customers) > 0 and len(services) > 0:
                # First valid sample
                sample_data.append({
                    'Customer': customers[0].name,
                    'Customer Reference No': 'REF001',
                    'Department/Person In Charge/Sub-Customer': 'Operations Department',
                    'Service': services[0].name,
                    'Pickup Date': (today + timedelta(days=1)).strftime('%d-%m-%Y'),
                    'Pickup Time': '09:00',
                    'Pickup Location': 'Sample Pickup Location 1',
                    'Drop-off Location': 'Sample Drop-off Location 1',
                    'Passenger Name': 'Sample Passenger 1',
                    'Passenger Mobile': '+6591234567',
                    'Vehicle Type': vehicle_type_value,
                    'Contractor': contractor_value,
                    'Remarks': 'Sample job entry - Valid data'
                })

                # Second valid sample
                sample_data.append({
                    'Customer': customers[0].name,
                    'Customer Reference No': 'REF002',
                    'Department/Person In Charge/Sub-Customer': 'Sales Department',
                    'Service': services[1].name if len(services) > 1 else services[0].name,
                    'Pickup Date': (today + timedelta(days=2)).strftime('%d-%m-%Y'),
                    'Pickup Time': '14:30',
                    'Pickup Location': 'Sample Pickup Location 2',
                    'Drop-off Location': 'Sample Drop-off Location 2',
                    'Passenger Name': 'Sample Passenger 2',
                    'Passenger Mobile': '+6598765432',
                    'Vehicle Type': vehicle_type_value_2,
                    'Contractor': contractor_value,
                    'Remarks': 'Sample job entry - Valid data'
                })
        elif customers and services and vehicles and drivers:
            # For Admin/Manager users: Original sample data with vehicles and drivers
            # First valid sample
            sample_data.append({
                'Customer': customers[0].name,
                'Customer Reference No': 'REF001',
                'Department/Person In Charge/Sub-Customer': 'Operations Department',
                'Service': services[0].name,
                'Pickup Date': (today + timedelta(days=1)).strftime('%d-%m-%Y'),
                'Pickup Time': '09:00',
                'Pickup Location': 'Sample Pickup Location 1',
                'Drop-off Location': 'Sample Drop-off Location 1',
                'Passenger Name': 'Sample Passenger 1',
                'Passenger Mobile': '+6591234567',
                'Vehicle Type': vehicle_type_value,
                'Vehicle': vehicles[0].number,
                'Driver': drivers[0].name,
                'Contractor': contractor_value,
                'Remarks': 'Sample job entry - Valid data'
            })

            # Second valid sample (if we have multiple records)
            if len(customers) > 1 and len(services) > 1 and len(vehicles) > 1 and len(drivers) > 1:
                sample_data.append({
                    'Customer': customers[1].name,
                    'Customer Reference No': 'REF002',
                    'Department/Person In Charge/Sub-Customer': 'Sales Department',
                    'Service': services[1].name if len(services) > 1 else services[0].name,
                    'Pickup Date': (today + timedelta(days=2)).strftime('%d-%m-%Y'),
                    'Pickup Time': '14:30',
                    'Pickup Location': 'Sample Pickup Location 2',
                    'Drop-off Location': 'Sample Drop-off Location 2',
                    'Passenger Name': 'Sample Passenger 2',
                    'Passenger Mobile': '+6598765432',
                    'Vehicle Type': vehicle_type_value_2,
                    'Vehicle': vehicles[1].number if len(vehicles) > 1 else vehicles[0].number,
                    'Driver': drivers[1].name if len(drivers) > 1 else drivers[0].name,
                    'Contractor': contractor_value_2,
                    'Remarks': 'Sample job entry - Valid data'
                })

            # Add invalid data samples for testing validation
            sample_data.append({
                'Customer': 'Invalid Customer',
                'Customer Reference No': 'REF003',
                'Department/Person In Charge/Sub-Customer': 'IT Department',
                'Service': services[0].name,
                'Pickup Date': (today + timedelta(days=3)).strftime('%d-%m-%Y'),
                'Pickup Time': '10:00',
                'Pickup Location': 'Test Location',
                'Drop-off Location': 'Test Destination',
                'Passenger Name': 'Test Passenger',
                'Passenger Mobile': '+6512345678',
                'Vehicle Type': vehicle_type_value,
                'Vehicle': vehicles[0].number,
                'Driver': drivers[0].name,
                'Contractor': contractor_value,
                'Remarks': 'Invalid customer - should fail validation'
            })

            sample_data.append({
                'Customer': customers[0].name,
                'Customer Reference No': 'REF004',
                'Department/Person In Charge/Sub-Customer': 'Marketing',
                'Service': 'Invalid Service',
                'Pickup Date': (today + timedelta(days=4)).strftime('%d-%m-%Y'),
                'Pickup Time': '11:00',
                'Pickup Location': 'Test Location',
                'Drop-off Location': 'Test Destination',
                'Passenger Name': 'Test Passenger',
                'Passenger Mobile': '+6587654321',
                'Vehicle Type': vehicle_type_value,
                'Vehicle': vehicles[0].number,
                'Driver': drivers[0].name,
                'Contractor': contractor_value,
                'Remarks': 'Invalid service - should fail validation'
            })

            sample_data.append({
                'Customer': customers[0].name,
                'Customer Reference No': 'REF005',
                'Department/Person In Charge/Sub-Customer': 'Finance',
                'Service': services[0].name,
                'Pickup Date': (today + timedelta(days=5)).strftime('%d-%m-%Y'),
                'Pickup Time': '12:00',
                'Pickup Location': 'Test Location',
                'Drop-off Location': 'Test Destination',
                'Passenger Name': 'Test Passenger',
                'Passenger Mobile': '+6596543210',
                'Vehicle Type': vehicle_type_value,
                'Vehicle': 'INVALID123',
                'Driver': drivers[0].name,
                'Contractor': contractor_value,
                'Remarks': 'Invalid vehicle - should fail validation'
            })

            sample_data.append({
                'Customer': customers[0].name,
                'Customer Reference No': 'REF006',
                'Department/Person In Charge/Sub-Customer': 'HR Department',
                'Service': services[0].name,
                'Pickup Date': (today + timedelta(days=6)).strftime('%d-%m-%Y'),
                'Pickup Time': '13:00',
                'Pickup Location': 'Test Location',
                'Drop-off Location': 'Test Destination',
                'Passenger Name': 'Test Passenger',
                'Passenger Mobile': '+6511223344',
                'Vehicle Type': vehicle_type_value,
                'Vehicle': vehicles[0].number,
                'Driver': 'Invalid Driver',
                'Contractor': contractor_value,
                'Remarks': 'Invalid driver - should fail validation'
            })

        # Fallback: Create empty template if no sample data was created
        if not sample_data:
            if is_customer_user:
                # Customer user template - NO Vehicle/Driver columns
                sample_data.append({
                    'Customer': '',
                    'Customer Reference No': '',
                    'Department/Person In Charge/Sub-Customer': '',
                    'Service': '',
                    'Pickup Date': (today + timedelta(days=1)).strftime('%d-%m-%Y'),
                    'Pickup Time': '09:00',
                    'Pickup Location': '',
                    'Drop-off Location': '',
                    'Passenger Name': '',
                    'Passenger Mobile': '',
                    'Vehicle Type': '',
                    'Contractor': '',
                    'Remarks': ''
                })
            else:
                # Admin user template - WITH Vehicle/Driver columns
                sample_data.append({
                    'Customer': '',
                    'Customer Reference No': '',
                    'Department/Person In Charge/Sub-Customer': '',
                    'Service': '',
                    'Pickup Date': (today + timedelta(days=1)).strftime('%d-%m-%Y'),
                    'Pickup Time': '09:00',
                    'Pickup Location': '',
                    'Drop-off Location': '',
                    'Passenger Name': '',
                    'Passenger Mobile': '',
                    'Vehicle Type': '',
                    'Vehicle': '',
                    'Driver': '',
                    'Contractor': '',
                    'Remarks': ''
                })
        
        # Create DataFrame
        df = pd.DataFrame(sample_data)
        
        # Create Excel file in memory using openpyxl engine
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Jobs Template', index=False)
            
            # Get the workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets['Jobs Template']
            
            # Style the header row
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
            
            # Header styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Apply header styling
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment

            # Create a hidden sheet for dropdown reference data
            # This solves the Excel 255-character limit for data validation formulas
            ref_sheet = workbook.create_sheet('ReferenceData')
            ref_sheet.sheet_state = 'hidden'

            # Add data validation for dropdowns using hidden reference sheet
            from openpyxl.worksheet.datavalidation import DataValidation

            # Customer dropdown (Column A) - Both customer and admin users
            if customers and len(customers) > 0:
                customer_names = [customer.name for customer in customers]
                # Write customer names to hidden sheet
                for idx, name in enumerate(customer_names, start=1):
                    ref_sheet.cell(row=idx, column=1, value=name)
                # Create named range reference
                customer_range = f"ReferenceData!$A$1:$A${len(customer_names)}"
                customer_validation = DataValidation(type="list", formula1=customer_range, allow_blank=True)
                customer_validation.add('A2:A1000')
                worksheet.add_data_validation(customer_validation)

            # Column B: Customer Reference No (text field - NO dropdown)
            # Column C: Department/Person In Charge/Sub-Customer (text field - NO dropdown)

            # Service dropdown (Column D) - Both customer and admin users
            if services and len(services) > 0:
                service_names = [service.name for service in services]
                # Write service names to hidden sheet (Column B)
                for idx, name in enumerate(service_names, start=1):
                    ref_sheet.cell(row=idx, column=2, value=name)
                service_range = f"ReferenceData!$B$1:$B${len(service_names)}"
                service_validation = DataValidation(type="list", formula1=service_range, allow_blank=True)
                service_validation.add('D2:D1000')
                worksheet.add_data_validation(service_validation)

            # Use non-overlapping columns for all reference data to prevent collision
            # Column A: Customers (all users)
            # Column B: Services (all users)
            # Column C: Vehicles (admin only)
            # Column D: Drivers (admin only)
            # Column E: Contractors (all users)
            # Column F: Vehicle Types (all users)

            # Vehicle Type dropdown (Column F in reference sheet) - All users
            # Column K for both customer and admin users (moved before Vehicle)
            if vehicle_types and len(vehicle_types) > 0:
                vehicle_type_names = [vtype.name for vtype in vehicle_types]
                # Write vehicle type names to hidden sheet (Column F)
                for idx, name in enumerate(vehicle_type_names, start=1):
                    ref_sheet.cell(row=idx, column=6, value=name)
                vehicle_type_range = f"ReferenceData!$F$1:$F${len(vehicle_type_names)}"
                vehicle_type_validation = DataValidation(type="list", formula1=vehicle_type_range, allow_blank=True)
                vehicle_type_validation.add('K2:K1000')  # Column K for both users
                worksheet.add_data_validation(vehicle_type_validation)

            # Vehicle dropdown (Column C in reference sheet) - Admin users only
            # Column L in main sheet (after Vehicle Type)
            if not is_customer_user and vehicles and len(vehicles) > 0:
                vehicle_numbers = [vehicle.number for vehicle in vehicles]
                # Write vehicle numbers to hidden sheet (Column C)
                for idx, number in enumerate(vehicle_numbers, start=1):
                    ref_sheet.cell(row=idx, column=3, value=number)
                vehicle_range = f"ReferenceData!$C$1:$C${len(vehicle_numbers)}"
                vehicle_validation = DataValidation(type="list", formula1=vehicle_range, allow_blank=True)
                vehicle_validation.add('L2:L1000')  # Column L in main sheet
                worksheet.add_data_validation(vehicle_validation)

            # Driver dropdown (Column D in reference sheet) - Admin users only
            # Column M in main sheet (after Vehicle)
            if not is_customer_user and drivers and len(drivers) > 0:
                driver_names = [driver.name for driver in drivers]
                # Write driver names to hidden sheet (Column D)
                for idx, name in enumerate(driver_names, start=1):
                    ref_sheet.cell(row=idx, column=4, value=name)
                driver_range = f"ReferenceData!$D$1:$D${len(driver_names)}"
                driver_validation = DataValidation(type="list", formula1=driver_range, allow_blank=True)
                driver_validation.add('M2:M1000')  # Column M in main sheet
                worksheet.add_data_validation(driver_validation)

            # Contractor dropdown (Column E in reference sheet) - All users
            # Column L for customer users, Column N for admin users (after driver details)
            if contractors and len(contractors) > 0:
                contractor_names = [contractor.name for contractor in contractors]
                # Write contractor names to hidden sheet (Column E)
                for idx, name in enumerate(contractor_names, start=1):
                    ref_sheet.cell(row=idx, column=5, value=name)
                contractor_range = f"ReferenceData!$E$1:$E${len(contractor_names)}"
                contractor_validation = DataValidation(type="list", formula1=contractor_range, allow_blank=True)
                # Column L for customer users, Column N for admin users
                target_column = 'L2:L1000' if is_customer_user else 'N2:N1000'
                contractor_validation.add(target_column)
                worksheet.add_data_validation(contractor_validation)

            # Auto-adjust column widths
            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Add instructions sheet - Customized based on user role
            if is_customer_user:
                # Customer user instructions - No Vehicle/Driver fields
                instructions_data = [
                    ['Instructions for Bulk Job Upload'],
                    [''],
                    ['1. Download this template'],
                    ['2. Fill in your job data (sample data is provided)'],
                    ['3. Save the file'],
                    ['4. Upload the completed file'],
                    [''],
                    ['Required Fields:'],
                    ['- Customer: Select from dropdown'],
                    ['- Service: Select from dropdown'],
                    ['- Pickup Date: Format DD-MM-YYYY or DDMMYYYY'],
                    ['- Pickup Time: Format HH:MM (24-hour)'],
                    ['- Pickup Location: Text'],
                    ['- Drop-off Location: Text'],
                    [''],
                    ['Optional Fields:'],
                    ['- Customer Reference No: Text'],
                    ['- Department/Person In Charge/Sub-Customer: Text'],
                    ['- Contractor: Select from dropdown'],
                    ['- Vehicle Type: Select from dropdown'],
                    ['- Passenger Name: Text'],
                    ['- Passenger Mobile: Text (with country code, e.g., +6591234567)'],
                    ['- Remarks: Text'],
                    [''],
                    ['Notes:'],
                    ['- Date format: DD-MM-YYYY (e.g., 25-01-2025) or DDMMYYYY (e.g., 25012025)'],
                    ['- Time format: HHMM (e.g., 0930, 1430) or HH:MM (e.g., 09:30, 14:30) in 24-hour format'],
                    ['- For times after midnight, use 4 digits: 0010 for 00:10, 0030 for 00:30, etc.'],
                    ['- Vehicle and Driver will be assigned by the system'],
                    ['- Use dropdowns for Customer, Service, Contractor, and Vehicle Type'],
                    ['- Remove sample data before uploading'],
                    ['- Maximum 1000 jobs per file'],
                    [''],
                    ['Job Status Rules:'],
                    ['- All uploaded jobs will be created with PENDING status'],
                    ['- Vehicle and Driver will be assigned later by administrators']
                ]
            else:
                # Admin user instructions - With Vehicle/Driver fields
                instructions_data = [
                    ['Instructions for Bulk Job Upload'],
                    [''],
                    ['1. Download this template'],
                    ['2. Fill in your job data (sample data is provided)'],
                    ['3. Save the file'],
                    ['4. Upload the completed file'],
                    [''],
                    ['Required Fields:'],
                    ['- Customer: Select from dropdown'],
                    ['- Service: Select from dropdown'],
                    ['- Vehicle: Select from dropdown'],
                    ['- Driver: Select from dropdown'],
                    ['- Pickup Date: Format DD-MM-YYYY or DDMMYYYY'],
                    ['- Pickup Time: Format HH:MM (24-hour)'],
                    ['- Pickup Location: Text'],
                    ['- Drop-off Location: Text'],
                    [''],
                    ['Optional Fields:'],
                    ['- Customer Reference No: Text'],
                    ['- Department/Person In Charge/Sub-Customer: Text'],
                    ['- Contractor: Select from dropdown'],
                    ['- Vehicle Type: Select from dropdown'],
                    ['- Passenger Name: Text'],
                    ['- Passenger Mobile: Text (with country code, e.g., +6591234567)'],
                    ['- Remarks: Text'],
                    [''],
                    ['Notes:'],
                    ['- Date format: DD-MM-YYYY (e.g., 25-01-2025) or DDMMYYYY (e.g., 25012025)'],
                    ['- Time format: HHMM (e.g., 0930, 1430) or HH:MM (e.g., 09:30, 14:30) in 24-hour format'],
                    ['- For times after midnight, use 4 digits: 0010 for 00:10, 0030 for 00:30, etc.'],
                    ['- Use dropdowns for Customer, Service, Vehicle, Driver, Contractor, and Vehicle Type'],
                    ['- Remove sample data before uploading'],
                    ['- Maximum 1000 jobs per file'],
                    [''],
                    ['Job Status Rules:'],
                    ['- Status will be automatically set based on provided fields:'],
                    ['  * CONFIRMED: When both Driver and Vehicle are assigned'],
                    ['  * PENDING: When only mandatory fields are filled (without Driver/Vehicle)'],
                    ['  * NEW: When mandatory fields are missing']
                ]
            
            instructions_df = pd.DataFrame(instructions_data)
            instructions_df.to_excel(writer, sheet_name='Instructions', index=False, header=False)
            
            # Style instructions sheet
            instructions_worksheet = writer.sheets['Instructions']
            instructions_worksheet.column_dimensions['A'].width = 60
        
        output.seek(0)
        
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=f'bulk_jobs_template_{datetime.now().strftime("%Y%m%d")}.xlsx'
        )
        
    except Exception as e:
        logging.error(f"Error generating Excel template: {str(e)}")
        return jsonify({'error': 'Error generating template. Please try again.'}), 500 

@job_bp.route('/jobs/upload', methods=['POST'])
@auth_required()
def upload_excel_file():
    """Handle Excel file upload and processing"""
    try:
        # Check if file was uploaded
        if 'file' not in request.files:
            return jsonify({'error': 'No file selected'}), 400
        
        file = request.files['file']
        
        # Check if file is empty
        if file.filename == '':
            return jsonify({'error': 'No file selected'}), 400
            
        # Check if file.filename is not None before using it
        if file.filename is None:
            return jsonify({'error': 'No file selected'}), 400
        
        # Check file extension
        allowed_extensions = current_app.config.get('ALLOWED_FILE_EXTENSIONS', {'.xlsx', '.xls'})
        file_ext = os.path.splitext(file.filename.lower())[1]
        if file_ext not in allowed_extensions:
            return jsonify({'error': f'Invalid file type. Please upload an Excel file ({", ".join(allowed_extensions)})'}), 400
        
        # Check file size
        file.seek(0, 2)  # Seek to end
        file_size = file.tell()
        file.seek(0)  # Reset to beginning
        
        max_file_size = current_app.config.get('MAX_FILE_SIZE', 10 * 1024 * 1024)
        if file_size > max_file_size:
            max_size_mb = max_file_size // (1024 * 1024)
            return jsonify({'error': f'File size too large. Please upload a file smaller than {max_size_mb}MB'}), 400
        
        if file_size == 0:
            return jsonify({'error': 'File is empty'}), 400
        
        # Create a secure filename
        filename = secure_filename(file.filename) if file.filename else 'upload.xlsx'

        # Capture user_role parameter from form data
        user_role = request.form.get('user_role', '')
        # Determine if user is a customer based on role or current_user
        is_customer_user = user_role == 'customer' or current_user.has_role('customer')

        # Save to temporary file with robust cleanup
        temp_file_path = None
        try:
            with tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx') as temp_file:
                file.save(temp_file.name)
                temp_file_path = temp_file.name


            # Process the Excel file for preview with user role
            preview_data = process_excel_file_preview(temp_file_path, is_customer_user=is_customer_user)
            
            return jsonify(preview_data)
                
        except Exception as e:
            logging.error(f"Error processing Excel file: {str(e)}")
            raise e
        finally:
            # Always clean up temporary file, even if an exception occurs
            if temp_file_path and os.path.exists(temp_file_path):
                try:
                    os.unlink(temp_file_path)

                except OSError as cleanup_error:
                    pass
        
    except Exception as e:
        logging.error(f"Error handling Excel upload: {str(e)}")
        return jsonify({'error': 'An error occurred while processing the file. Please try again.'}), 500


def process_excel_file_preview(file_path, column_mapping=None, is_customer_user=False):
    """Process the uploaded Excel file for preview with validation and column mapping"""
    try:
        # Read Excel file with pandas
        df = pd.read_excel(file_path, sheet_name='Jobs Template')

        # Debug logging to diagnose empty rows issue
        current_app.logger.info(f"Excel file loaded: {len(df)} rows found (excluding header)")
        current_app.logger.info(f"DataFrame shape: {df.shape}")
        current_app.logger.info(f"DataFrame columns: {df.columns.tolist()}")
        if len(df) > 0:
            current_app.logger.info(f"First row data: {df.iloc[0].to_dict()}")

        # Use openpyxl to get raw cell values for time columns to preserve leading zeros
        wb = None
        try:
            wb = load_workbook(file_path, data_only=False)
            ws = wb['Jobs Template']

            # Find the time column index
            header_row = [cell.value for cell in ws[1]]
            time_column_idx = None
            time_column_names = ['Pickup Time', 'Time', 'pickup_time']
            for idx, col_name in enumerate(header_row):
                if col_name in time_column_names:
                    time_column_idx = idx + 1  # openpyxl is 1-indexed
                    break

            # If we found a time column, replace pandas values with raw cell values from openpyxl
            if time_column_idx:
                time_values = []
                for row_idx in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
                    cell = ws.cell(row=row_idx, column=time_column_idx)

                    # Check if cell is formatted as text or if it's a string value
                    if cell.value is not None:
                        # If it's already a string, use it directly
                        if isinstance(cell.value, str):
                            raw_value = cell.value
                        # If it's a number, check if it looks like it should have leading zeros
                        elif isinstance(cell.value, (int, float)):
                            # For numbers less than 100, assume they might be times like "0010" or "0930"
                            # We'll pad them to 4 digits
                            num_str = str(int(cell.value))
                            if int(cell.value) < 100:
                                # Could be "0010" (00:10) or "0930" (09:30)
                                raw_value = num_str.zfill(4)
                            else:
                                raw_value = num_str
                        else:
                            raw_value = str(cell.value)
                    else:
                        raw_value = ''

                    time_values.append(raw_value)

                # Update the dataframe with raw time values
                time_col_name = header_row[time_column_idx - 1]
                if len(time_values) == len(df):
                    df[time_col_name] = time_values
                else:
                    current_app.logger.warning(
                        f"Time column row mismatch: openpyxl={len(time_values)}, pandas={len(df)}. "
                        "Using pandas values; leading zeros in times may be lost."
                    )
        finally:
            if wb:
                wb.close()
        
        # Check row count limit
        max_rows = current_app.config.get('MAX_ROWS_PER_FILE', 1000)
        if len(df) > max_rows:
            return {
                'valid_count': 0,
                'error_count': 1,
                'rows': [{
                    'row_number': 1,
                    'customer': 'N/A',
                    'customer_reference_no': 'N/A',
                    'department': 'N/A',
                    'service': 'N/A',
                    'vehicle': 'N/A',
                    'driver': 'N/A',
                    'contractor': 'N/A',
                    'vehicle_type': 'N/A',
                    'pickup_date': 'N/A',
                    'pickup_time': 'N/A',
                    'pickup_location': 'N/A',
                    'dropoff_location': 'N/A',
                    'passenger_name': 'N/A',
                    'passenger_mobile': 'N/A',
                    'remarks': 'N/A',
                    'is_valid': False,
                    'error_message': f"File contains too many rows ({len(df)}). Maximum allowed is {max_rows} rows."
                }],
                'json_data': json.dumps([]),
                'column_mapping': {},
                'available_columns': []
            }
        
        # Default column mapping
        default_mapping = {
            'customer': ['Customer', 'Agent', 'Client'],
            'customer_reference_no': ['Customer Reference No', 'Reference No', 'Ref No', 'Customer Ref'],
            'department': ['Department/Person In Charge/Sub-Customer', 'Department', 'Person In Charge', 'Sub-Customer'],
            'service': ['Service', 'Type', 'Category'],
            'vehicle': ['Vehicle', 'Car', 'Vehicle Number'],
            'driver': ['Driver', 'Chauffeur', 'Driver Name'],
            'contractor': ['Contractor', 'Contractor Name', 'Vendor'],
            'vehicle_type': ['Vehicle Type', 'Type of Vehicle', 'Car Type'],
            'pickup_date': ['Pickup Date', 'Date', 'Pickup Date'],
            'pickup_time': ['Pickup Time', 'Time', 'Pickup Time'],
            'pickup_location': ['Pickup Location', 'From', 'Pickup Location'],
            'dropoff_location': ['Drop-off Location', 'To', 'Drop-off Location'],
            'passenger_name': ['Passenger Name', 'Passenger', 'Name'],
            'passenger_mobile': ['Passenger Mobile', 'Mobile', 'Phone', 'Contact Number'],
            'remarks': ['Remarks', 'Notes', 'Comments']
        }
        
        # Use provided mapping or default
        mapping = column_mapping or default_mapping
        
        # Find actual column names in the file
        actual_columns = df.columns.tolist()
        
        # Map columns
        column_map = {}
        for field, possible_names in mapping.items():
            for name in possible_names:
                if name in actual_columns:
                    column_map[field] = name
                    break
            if field not in column_map:
                column_map[field] = possible_names[0]  # Use first as default
        
        # Validate required columns based on user role
        if is_customer_user:
            # Customer users: Vehicle and Driver are NOT required
            required_fields = ['customer', 'service', 'pickup_date', 'pickup_time',
                              'pickup_location', 'dropoff_location']
        else:
            # Admin users: All fields including Vehicle and Driver are required
            required_fields = ['customer', 'service', 'vehicle', 'driver', 'pickup_date', 'pickup_time',
                              'pickup_location', 'dropoff_location']

        missing_columns = []
        for field in required_fields:
            if field not in column_map or column_map[field] not in actual_columns:
                missing_columns.append(field)
        
        if missing_columns:
            return {
                'valid_count': 0,
                'error_count': 1,
                'rows': [{
                    'row_number': 1,
                    'customer': 'N/A',
                    'customer_reference_no': 'N/A',
                    'department': 'N/A',
                    'service': 'N/A',
                    'vehicle': 'N/A',
                    'driver': 'N/A',
                    'contractor': 'N/A',
                    'vehicle_type': 'N/A',
                    'pickup_date': 'N/A',
                    'pickup_time': 'N/A',
                    'pickup_location': 'N/A',
                    'dropoff_location': 'N/A',
                    'passenger_name': 'N/A',
                    'passenger_mobile': 'N/A',
                    'remarks': 'N/A',
                    'is_valid': False,
                    'error_message': f"Missing required columns: {', '.join(missing_columns)}"
                }],
                'json_data': json.dumps([]),
                'column_mapping': column_map,
                'available_columns': actual_columns
            }
        
        # Get lookup data using centralized validation
        lookups = get_validation_lookups()
        preview_rows = []
        valid_count = 0
        error_count = 0

        # Helper function to normalize time format (defined outside loop for performance)
        def normalize_time(time_str):
            """Convert time without colon (e.g., '0900', '930') to HH:MM format"""
            if not time_str:
                return ''

            # Convert to string and strip whitespace
            original_str = str(time_str).strip()

            # If already has colon, return as is
            if ':' in original_str:
                return original_str

            # Remove any non-digit characters (including decimal points from float conversion)
            digits = ''.join(c for c in original_str if c.isdigit())

            if not digits:
                return original_str

            # Handle different formats based on digit length
            if len(digits) == 4:
                # e.g., "0930" -> "09:30", "0010" -> "00:10"
                return f"{digits[0:2]}:{digits[2:4]}"
            elif len(digits) == 3:
                # e.g., "930" -> "09:30"
                return f"{digits[0]:0>2}:{digits[1:3]}"
            elif len(digits) == 1 or len(digits) == 2:
                # e.g., "9" -> "09:00", "10" -> "10:00"
                # Check if original string had leading zeros (preserved by openpyxl)
                if original_str.startswith('0') and len(original_str) >= 3:
                    # This was "0010" or similar - pad to 4 digits
                    digits = digits.zfill(4)
                    return f"{digits[0:2]}:{digits[2:4]}"
                else:
                    return f"{digits:0>2}:00"

            return original_str

        # Helper function to clean and normalize cell values (defined outside loop for performance)
        def clean_value(value, is_date=False, is_time=False):
            if pd.isna(value):
                return ''
            # Handle datetime objects from Excel
            if isinstance(value, pd.Timestamp):
                if is_date:
                    return value.strftime('%d-%m-%Y')
                return value.strftime('%H:%M')
            # Handle Python datetime objects
            if hasattr(value, 'strftime'):
                if is_date:
                    return value.strftime('%d-%m-%Y')
                return value.strftime('%H:%M')

            result = str(value).strip()

            # Handle string "nan" from pandas when dtype is forced to string
            if result.lower() == 'nan':
                return ''

            # Normalize time format if this is a time field
            if is_time:
                result = normalize_time(result)

            return result

        # Process each row
        for index, row in df.iterrows():
            # Get row index as integer
            row_index = index if isinstance(index, int) else 0

            # Conditional row data extraction based on user role
            if is_customer_user:
                # Customer users: Vehicle and Driver set to empty strings
                row_data = {
                    'row_number': row_index + 2,  # Excel rows start from 2 (1 is header)
                    'customer': clean_value(row.get(column_map.get('customer', 'Customer'), '')),
                    'customer_reference_no': clean_value(row.get(column_map.get('customer_reference_no', 'Customer Reference No'), '')),
                    'department': clean_value(row.get(column_map.get('department', 'Department/Person In Charge/Sub-Customer'), '')),
                    'service': clean_value(row.get(column_map.get('service', 'Service'), '')),
                    'vehicle': '',  # Always empty for customer users
                    'driver': '',   # Always empty for customer users
                    'contractor': clean_value(row.get(column_map.get('contractor', 'Contractor'), '')),
                    'vehicle_type': clean_value(row.get(column_map.get('vehicle_type', 'Vehicle Type'), '')),
                    'pickup_date': clean_value(row.get(column_map.get('pickup_date', 'Pickup Date'), ''), is_date=True),
                    'pickup_time': clean_value(row.get(column_map.get('pickup_time', 'Pickup Time'), ''), is_time=True),
                    'pickup_location': clean_value(row.get(column_map.get('pickup_location', 'Pickup Location'), '')),
                    'dropoff_location': clean_value(row.get(column_map.get('dropoff_location', 'Drop-off Location'), '')),
                    'passenger_name': clean_value(row.get(column_map.get('passenger_name', 'Passenger Name'), '')),
                    'passenger_mobile': clean_value(row.get(column_map.get('passenger_mobile', 'Passenger Mobile'), '')),
                    'remarks': clean_value(row.get(column_map.get('remarks', 'Remarks'), '')),
                    'is_valid': True,
                    'error_message': ''
                }
            else:
                # Admin users: Vehicle and Driver read from columns
                row_data = {
                    'row_number': row_index + 2,  # Excel rows start from 2 (1 is header)
                    'customer': clean_value(row.get(column_map.get('customer', 'Customer'), '')),
                    'customer_reference_no': clean_value(row.get(column_map.get('customer_reference_no', 'Customer Reference No'), '')),
                    'department': clean_value(row.get(column_map.get('department', 'Department/Person In Charge/Sub-Customer'), '')),
                    'service': clean_value(row.get(column_map.get('service', 'Service'), '')),
                    'vehicle': clean_value(row.get(column_map.get('vehicle', 'Vehicle'), '')),
                    'driver': clean_value(row.get(column_map.get('driver', 'Driver'), '')),
                    'contractor': clean_value(row.get(column_map.get('contractor', 'Contractor'), '')),
                    'vehicle_type': clean_value(row.get(column_map.get('vehicle_type', 'Vehicle Type'), '')),
                    'pickup_date': clean_value(row.get(column_map.get('pickup_date', 'Pickup Date'), ''), is_date=True),
                    'pickup_time': clean_value(row.get(column_map.get('pickup_time', 'Pickup Time'), ''), is_time=True),
                    'pickup_location': clean_value(row.get(column_map.get('pickup_location', 'Pickup Location'), '')),
                    'dropoff_location': clean_value(row.get(column_map.get('dropoff_location', 'Drop-off Location'), '')),
                    'passenger_name': clean_value(row.get(column_map.get('passenger_name', 'Passenger Name'), '')),
                    'passenger_mobile': clean_value(row.get(column_map.get('passenger_mobile', 'Passenger Mobile'), '')),
                    'remarks': clean_value(row.get(column_map.get('remarks', 'Remarks'), '')),
                    'is_valid': True,
                    'error_message': ''
                }

            # Validate mandatory fields first (before centralized validation)
            mandatory_fields = ['customer', 'service', 'pickup_date', 'pickup_time',
                              'pickup_location', 'dropoff_location', 'passenger_name']
            missing_fields = [f for f in mandatory_fields if not row_data.get(f, '').strip()]

            if missing_fields:
                row_data['is_valid'] = False
                # Format missing fields consistently with other validation errors
                # Use semicolon separator to allow frontend to split and display individually
                if len(missing_fields) == 1:
                    row_data['error_message'] = f"{missing_fields[0]} is required"
                else:
                    missing_field_errors = [f"{field} is required" for field in missing_fields]
                    row_data['error_message'] = '; '.join(missing_field_errors)
                error_count += 1
                preview_rows.append(row_data)
                continue

            # Validate date/time formats
            from datetime import datetime
            pickup_date_val = row_data.get('pickup_date', '').strip()
            pickup_time_val = row_data.get('pickup_time', '').strip()

            if pickup_date_val:
                # Parse DD-MM-YYYY, DDMMYYYY, or YYYY-MM-DD format and convert to YYYY-MM-DD for database
                date_obj = None
                for fmt in ('%d-%m-%Y', '%d%m%Y', '%Y-%m-%d'):
                    try:
                        date_obj = datetime.strptime(pickup_date_val, fmt)
                        break
                    except ValueError:
                        continue

                if date_obj:
                    row_data['pickup_date'] = date_obj.strftime('%Y-%m-%d')
                else:
                    row_data['is_valid'] = False
                    row_data['error_message'] = f"Invalid pickup_date format: '{pickup_date_val}'. Expected DD-MM-YYYY, DDMMYYYY, or YYYY-MM-DD"
                    error_count += 1
                    preview_rows.append(row_data)
                    continue

            if pickup_time_val:
                try:
                    # Normalize time format: handle both HH:MM and HH:MM:SS formats
                    if pickup_time_val.count(':') == 2:  # HH:MM:SS format
                        pickup_time_val = ':'.join(pickup_time_val.split(':')[:2])  # Strip seconds
                        row_data['pickup_time'] = pickup_time_val  # Update with normalized value
                    datetime.strptime(pickup_time_val, '%H:%M')
                except ValueError:
                    row_data['is_valid'] = False
                    row_data['error_message'] = f"Invalid pickup_time format: '{pickup_time_val}'. Expected HH:MM"
                    error_count += 1
                    preview_rows.append(row_data)
                    continue

            # Use centralized validation
            try:
                is_valid, error_message, validated_data = validate_job_row(row_data, lookups)
                
                # Update with validated data (including IDs) first
                row_data.update(validated_data)
                
                # Then update validation results (this should override any values from validated_data)
                row_data['is_valid'] = is_valid
                row_data['error_message'] = error_message
                
                if is_valid:
                    valid_count += 1
                else:
                    error_count += 1
                
                # Add row to preview_rows (this was missing!)
                preview_rows.append(row_data)
                    
            except Exception as e:
                # Fallback to original validation logic if centralized validation fails
                logging.error(f"Centralized validation failed, using fallback: {e}")
                
                # Skip empty rows - convert to scalar values for comparison
                customer_val = row.get(column_map.get('customer', 'Customer'))
                pickup_loc_val = row.get(column_map.get('pickup_location', 'Pickup Location'))
                
                # Check if values are NaN or empty using simple checks
                is_customer_empty = pd.isna(customer_val) or (isinstance(customer_val, str) and customer_val.strip() == '')
                is_pickup_empty = pd.isna(pickup_loc_val) or (isinstance(pickup_loc_val, str) and pickup_loc_val.strip() == '')
                
                # Use simple boolean conversion
                if bool(is_customer_empty) or bool(is_pickup_empty):
                    row_data['is_valid'] = False
                    row_data['error_message'] = 'Empty required fields'
                    error_count += 1
                    preview_rows.append(row_data)
                    continue
                
                # Validate customer
                if row_data['customer'] not in lookups['customers']:
                    row_data['is_valid'] = False
                    row_data['error_message'] = f"Customer '{row_data['customer']}' not found"
                    error_count += 1
                    preview_rows.append(row_data)
                    continue
                
                # Validate service
                if row_data['service'] not in lookups['services']:
                    row_data['is_valid'] = False
                    row_data['error_message'] = f"Service '{row_data['service']}' not found"
                    error_count += 1
                    preview_rows.append(row_data)
                    continue
                
                # If all validations pass
                row_data['is_valid'] = True
                row_data['error_message'] = ''
                valid_count += 1
                preview_rows.append(row_data)
        

        
        return {
            'valid_count': valid_count,
            'error_count': error_count,
            'rows': preview_rows,
            'json_data': json.dumps(preview_rows),
            'column_mapping': column_map,
            'available_columns': actual_columns
        }
        
    except Exception as e:
        return {
            'valid_count': 0,
            'error_count': 1,
            'rows': [{
                'row_number': 1,
                'customer': 'N/A',
                'customer_reference_no': 'N/A',
                'department': 'N/A',
                'service': 'N/A',
                'vehicle': 'N/A',
                'driver': 'N/A',
                'contractor': 'N/A',
                'vehicle_type': 'N/A',
                'pickup_date': 'N/A',
                'pickup_time': 'N/A',
                'pickup_location': 'N/A',
                'dropoff_location': 'N/A',
                'passenger_name': 'N/A',
                'passenger_mobile': 'N/A',
                'status': 'N/A',
                'remarks': 'N/A',
                'is_valid': False,
                'error_message': "Error reading Excel file. Please check the file format and try again."
            }],
            'json_data': json.dumps([]),
            'column_mapping': {},
            'available_columns': []
        }


@job_bp.route('/jobs/confirm-upload', methods=['POST'])
@roles_accepted('admin', 'manager', 'accountant', 'customer')
def confirm_upload():
    """Handle confirmation of preview data and create jobs"""
    try:
        # Check if user is a Customer role
        is_customer_user = current_user.has_role('customer')
        customer_user_customer_id = None

        if is_customer_user:
            # Get the customer_id for the logged-in customer user
            customer_user_customer_id = getattr(current_user, 'customer_id', None)
            if not customer_user_customer_id:
                return jsonify({'error': 'Customer user does not have a customer_id assigned'}), 403

        # Get the preview data from the request
        preview_data = request.get_json()
        if not preview_data or 'rows' not in preview_data:
            return jsonify({'error': 'No preview data found'}), 400

        # Get optional parameter to allow duplicate jobs (for recurring uploads)
        allow_duplicates = preview_data.get('allow_duplicates', False)

        if allow_duplicates:
            logging.info("Duplicate detection disabled for this upload (allow_duplicates=True)")

        # Get lookup data
        customers = {customer.name: customer for customer in Customer.query.filter_by(status='Active').all()}
        services = {service.name: service.name for service in Service.query.filter_by(status='Active').all()}
        vehicles = {vehicle.number: vehicle for vehicle in Vehicle.query.filter_by(status='Active').all()}
        drivers = {driver.name: driver for driver in Driver.query.filter_by(status='Active').all()}
        contractors = {contractor.name: contractor for contractor in Contractor.query.filter_by(status='Active').all()}
        vehicle_types = {vtype.name: vtype for vtype in VehicleType.query.filter_by(status=True, is_deleted=False).all()}
        
        # Debug logging


        # First pass: Validate all rows and collect validation errors
        invalid_rows = []
        valid_jobs_data = []

        for row_data in preview_data['rows']:
            if not row_data.get('is_valid', False):
                # Track invalid rows that were already marked invalid during preview
                invalid_rows.append({
                    'row': row_data.get('row_number', 'unknown'),
                    'customer': row_data.get('customer', ''),
                    'service': row_data.get('service', ''),
                    'errors': [row_data.get('error_message', 'Row is invalid')]
                })
                continue

            # Validate each row and collect errors
            row_errors = []
            row_index = row_data.get('row_number', 'unknown')

            try:
                # Get related objects
                customer = customers.get(row_data['customer'])
                vehicle = vehicles.get(row_data['vehicle'])
                driver = drivers.get(row_data['driver'])
                contractor = contractors.get(row_data.get('contractor', '')) if row_data.get('contractor') else None
                vehicle_type = vehicle_types.get(row_data.get('vehicle_type', '')) if row_data.get('vehicle_type') else None

                # Validate customer exists
                if not customer:
                    row_errors.append(f"Customer '{row_data['customer']}' not found or not active")

                # Get service by name
                service_name = row_data['service']
                service = Service.query.filter_by(name=service_name, status='Active').first()
                if not service:
                    row_errors.append(f"Service '{service_name}' not found or not active")

                # For customer users: Validate they can only create jobs for their own customer
                if is_customer_user and customer:
                    if customer.id != customer_user_customer_id:
                        row_errors.append(f"Customer users can only create jobs for their own customer")
                    # Validate contractor is AG Internal
                    if contractor and contractor.name not in ['AG', 'AG (Internal)']:
                        row_errors.append("Customer users can only assign jobs to AG Internal contractor")

                # Validate all mandatory fields exist and have non-empty values
                mandatory_fields = ['customer', 'service', 'pickup_date', 'pickup_time',
                                  'pickup_location', 'dropoff_location', 'passenger_name']
                for field in mandatory_fields:
                    value = str(row_data.get(field, '')).strip()
                    if not value:
                        row_errors.append(f"Missing or empty required field: {field}")

                # Validate and parse date/time formats
                pickup_date = str(row_data.get('pickup_date', '')).strip()
                pickup_time = str(row_data.get('pickup_time', '')).strip()

                if pickup_date:
                    try:
                        from datetime import datetime
                        # Parse DD-MM-YYYY, DDMMYYYY, or YYYY-MM-DD format and convert to YYYY-MM-DD for database
                        try:
                            date_obj = datetime.strptime(pickup_date, '%d-%m-%Y')
                        except ValueError:
                            try:
                                date_obj = datetime.strptime(pickup_date, '%d%m%Y')
                            except ValueError:
                                # Also accept YYYY-MM-DD format (from preview data)
                                date_obj = datetime.strptime(pickup_date, '%Y-%m-%d')
                        pickup_date = date_obj.strftime('%Y-%m-%d')
                    except ValueError:
                        row_errors.append(f"Invalid pickup_date format: '{pickup_date}'. Expected DD-MM-YYYY, DDMMYYYY, or YYYY-MM-DD")

                if pickup_time:
                    try:
                        from datetime import datetime
                        # Normalize time format: handle both HH:MM and HH:MM:SS formats
                        if pickup_time.count(':') == 2:  # HH:MM:SS format
                            pickup_time = ':'.join(pickup_time.split(':')[:2])  # Strip seconds
                        datetime.strptime(pickup_time, '%H:%M')
                    except ValueError:
                        row_errors.append(f"Invalid pickup_time format: '{pickup_time}'. Expected HH:MM (24-hour)")

                # If there are validation errors, skip this row
                if row_errors:
                    invalid_rows.append({
                        'row': row_index,
                        'customer': row_data.get('customer', ''),
                        'service': service_name,
                        'errors': row_errors
                    })
                    continue

                # Determine job status based on available fields
                mandatory_fields = ['customer', 'service', 'pickup_date', 'pickup_time',
                                  'pickup_location', 'dropoff_location', 'passenger_name']
                mandatory_filled = all(str(row_data.get(f, '')).strip() for f in mandatory_fields)

                job_status = 'new'
                if driver and vehicle and mandatory_filled:
                    job_status = 'confirmed'
                elif mandatory_filled:
                    job_status = 'pending'

                # Collect valid job data for batch processing
                job_data = {
                    'row_number': row_index,
                    'customer_id': customer.id,
                    'booking_ref': row_data.get('customer_reference_no', ''),
                    'sub_customer_name': row_data.get('department', ''),
                    'service_type': service.name,
                    'service_id': service.id,
                    'vehicle_id': vehicle.id if vehicle else None,
                    'driver_id': driver.id if driver else None,
                    'contractor_id': contractor.id if contractor else None,
                    'vehicle_type_id': vehicle_type.id if vehicle_type else None,
                    'pickup_location': str(row_data.get('pickup_location', '')).strip(),
                    'dropoff_location': str(row_data.get('dropoff_location', '')).strip(),
                    'pickup_date': pickup_date,
                    'pickup_time': pickup_time,
                    'passenger_name': str(row_data.get('passenger_name', '')).strip(),
                    'passenger_mobile': str(row_data.get('passenger_mobile', '')).strip(),
                    'status': job_status,
                    'customer_remark': row_data.get('remarks', ''),
                    'original_row_data': row_data  # Keep reference for duplicate check
                }
                valid_jobs_data.append(job_data)

            except Exception as validation_error:
                # Catch any unexpected validation errors
                logging.error(f"Validation error for row {row_index}: {str(validation_error)}")
                invalid_rows.append({
                    'row': row_index,
                    'customer': row_data.get('customer', ''),
                    'service': row_data.get('service', ''),
                    'errors': [str(validation_error)]
                })

        # Report validation failures before processing any jobs
        if invalid_rows:
            logging.warning(f"Bulk upload validation: {len(invalid_rows)} invalid rows out of {len(preview_data['rows'])}")
            return jsonify({
                'error': 'Some rows failed validation',
                'invalid_rows': invalid_rows,
                'valid_count': len(valid_jobs_data),
                'invalid_count': len(invalid_rows)
            }), 422

        # Second pass: Process all valid jobs
        processed_count = 0
        errors = []
        skipped_rows = []
        created_jobs = []

        for job_data in valid_jobs_data:
            row_number = job_data.pop('row_number')
            original_row_data = job_data.pop('original_row_data')

            try:
                # Check if job already exists (duplicate detection)
                if not allow_duplicates:
                    existing_job = Job.query.filter_by(
                        customer_id=job_data['customer_id'],
                        pickup_location=job_data['pickup_location'],
                        dropoff_location=job_data['dropoff_location'],
                        pickup_date=job_data['pickup_date'],
                        pickup_time=job_data['pickup_time'],
                        service_type=job_data['service_type'],
                        is_deleted=False
                    ).first()

                    if existing_job:
                        skipped_rows.append({
                            'row_number': row_number,
                            'reason': f'Duplicate job - already exists as Job #{existing_job.id}'
                        })
                        logging.warning(f"Skipping duplicate job for row {row_number}: Job #{existing_job.id}")
                        continue

                # Check for driver scheduling conflict
                if job_data.get('driver_id') and job_data.get('pickup_date') and job_data.get('pickup_time'):
                    conflict_job = JobService.check_driver_conflict(
                        job_data['driver_id'],
                        job_data['pickup_date'],
                        job_data['pickup_time']
                    )
                    if conflict_job:
                        if not current_user.has_role('admin'):
                            raise Exception(
                                f"Scheduling conflict for driver {job_data['driver_id']} "
                                f"at {job_data['pickup_date']} {job_data['pickup_time']}"
                            )
                        else:
                            logging.warning(
                                f"Bulk upload: Admin overriding conflict for driver {job_data['driver_id']} "
                                f"at {job_data['pickup_date']} {job_data['pickup_time']}. "
                                f"Conflict with job #{conflict_job.id}"
                            )

                # JobService.create handles its own commit/rollback
                job = JobService.create(job_data)
                processed_count += 1

                # Track created job details
                created_jobs.append({
                    'job_id': f"JOB-{job.id}",
                    'row_number': row_number,
                    'customer': original_row_data.get('customer', ''),
                    'pickup_date': original_row_data.get('pickup_date', '')
                })

            except Exception as row_error:
                # Job creation failed for this row, continue with next row
                error_msg = f"Error processing row {row_number}: {str(row_error)}"
                logging.error(error_msg, exc_info=True)
                errors.append(error_msg)

        return jsonify({
            'processed_count': processed_count,
            'skipped_count': len(skipped_rows),
            'skipped_rows': skipped_rows,
            'errors': errors,
            'created_jobs': created_jobs
        }), 200

    except Exception as e:
        logging.error(f"Error handling bulk upload confirmation: {str(e)}", exc_info=True)
        return jsonify({'error': 'An error occurred while processing the bulk upload. Please try again.'}), 500

@job_bp.route('/jobs/audit/<int:job_id>', methods=['POST'])
@roles_accepted('admin', 'manager', 'driver', 'customer', 'accountant')
def create_job_audit(job_id):
    """Create a job audit record for status changes or other job modifications."""
    try:
        # Get the job to ensure it exists with row-level locking
        job = Job.query.with_for_update().get(job_id)
        if not job:
            return jsonify({'error': 'Job not found'}), 404
            
        # Get JSON data from request
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
            
        # Validate required fields
        new_status = data.get('new_status')
        if not new_status:
            return jsonify({'error': 'new_status is required'}), 400
            
        # Validate that the new status is a valid job status
        valid_statuses = [status.value for status in JobStatus]
        if new_status not in valid_statuses:
            return jsonify({'error': f'Invalid status. Must be one of: {", ".join(valid_statuses)}'}), 400
            
        # Capture the current job status before updating it
        old_status = job.status
        
        # Validate reason for cancellation
        reason = data.get('reason')
        if new_status == JobStatus.CANCELED.value and not reason:
            return jsonify({'error': 'reason is required for cancellation'}), 400
        
        # Create audit record
        audit_record = JobAudit(
            job_id=job_id,
            changed_by=current_user.id if current_user.is_authenticated else None,
            old_status=old_status,
            new_status=new_status,
            reason=data.get('reason'),
            additional_data=data.get('additional_data')
        )
        
        # Add to database
        db.session.add(audit_record)
        
        # Update job status if provided
        if new_status:
            job.status = new_status
            
        # Commit both changes atomically
        db.session.commit()
        
        return jsonify({
            'message': 'Job audit record created successfully with row-level locking',
            'id': audit_record.id
        }), 201
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error creating job audit record: {e}", exc_info=True)
        return jsonify({'error': 'An error occurred while creating the audit record'}), 500


@job_bp.route('/jobs/audit_trail/<int:job_id>', methods=['GET'])
@roles_accepted('admin', 'manager', 'customer', 'driver', 'accountant')
def get_job_audit_records(job_id):
    """Get all audit records for a specific job."""
    try:
        # Get the job to ensure it exists and include driver information
        job = Job.query.options(db.joinedload(Job.driver)).get(job_id)
        if not job:
            return jsonify({'error': 'Job not found'}), 404
            
        # Permissions are already checked by @roles_accepted('admin', 'manager') decorator
            
        # Get audit records ordered by changed_at descending (newest first)
        audit_records = (
            JobAudit.query
            .options(db.joinedload(JobAudit.changed_by_user))
            .filter_by(job_id=job_id)
            .order_by(JobAudit.changed_at.desc())
            .all()
        )
        
        # Convert to JSON-friendly format
        audit_data = []
        for record in audit_records:
            # Get user name - use email if no name field exists, or driver name if user is linked to driver
            changed_by_name = record.changed_by_user.email if record.changed_by_user else None
            if record.changed_by_user and record.changed_by_user.driver:
                changed_by_name = record.changed_by_user.driver.name
            
            # Get user role - get the first role name if user has roles
            role = None
            if record.changed_by_user and record.changed_by_user.roles:
                role = record.changed_by_user.roles[0].name if record.changed_by_user.roles else None
            
            # Create status labels - for now just use the status values as labels
            old_status_label = record.old_status
            new_status_label = record.new_status
            
            # Get remarks based on status transition
            remark = None
            extra_services = None  # Initialize extra_services variable
            remarks_list = []  
            if (record.new_status == 'confirmed' and record.old_status == 'new') or \
               (record.new_status == 'confirmed' and record.old_status == 'pending'):
                # Show customer remark when transitioning from new/pending to confirmed
                remark = job.customer_remark
                
                # Also include extra services when status is confirmed
                if job.extra_services_data:
                    extra_services_list = []
                    try:
                        if isinstance(job.extra_services_data, list):
                            for service in job.extra_services_data:
                                if isinstance(service, dict):
                                    desc = service.get('description') or service.get('name')
                                    if desc and isinstance(desc, str):
                                        extra_services_list.append(desc)
                                elif isinstance(service, str):
                                    extra_services_list.append(service)
                        if extra_services_list:
                            extra_services = extra_services_list
                    except Exception as e:
                        # Log error but don't fail the entire request
                        current_app.logger.error(f"Error parsing extra_services for job {job_id}: {str(e)}")
                        extra_services = None
            else:
                # Get driver remark for non-confirmation transitions
                if record.changed_at:
                    changed_time = record.changed_at.replace(tzinfo=None) if record.changed_at.tzinfo else record.changed_at
           
                    closest_driver_remark = (
                        DriverRemark.query
                        .filter(DriverRemark.job_id == job_id)
                        .order_by(
                            func.abs(
                                func.strftime('%s', DriverRemark.created_at) -
                                func.strftime('%s', changed_time)
                            )
                        )
                        .first()
                    )

                    if closest_driver_remark:
                        remark = closest_driver_remark.remark
                        remarks_list.append({
                            'id': closest_driver_remark.id,
                            'remark': closest_driver_remark.remark,
                            'created_at': closest_driver_remark.created_at.isoformat()
                                          if closest_driver_remark.created_at else None
                        })
            # Get photos uploaded for the same stage as this audit record
            attachments = []
            if record.changed_at and record.new_status:
                # Map job status to photo stage
                # Photos are stored with the actual status as stage, not generic stages
                status_to_photo_stage = {
                    'OTW': 'OTW',
                    'POB': 'POB',
                    'JC': 'JC',
                    'SD': 'SD'
                }
                
                # Get the photo stage based on the new status
                photo_stage = status_to_photo_stage.get(record.new_status.upper())
                
                if photo_stage:
                    # Get photos for the specific stage that were uploaded within a reasonable time window
                    # around the audit event to increase likelihood of successful photo attachment
                    # Using a 30-minute window (15 minutes before and 15 minutes after) to be more inclusive
                    
                    # Handle potential timezone differences between audit record and photo timestamps
                    changed_at_utc = record.changed_at
                    if changed_at_utc.tzinfo is not None:
                        # If audit record has timezone info, convert to naive datetime for comparison
                        changed_at_utc = changed_at_utc.replace(tzinfo=None)
                    
                    photo_query = JobPhoto.query.filter(
                        JobPhoto.job_id == record.job_id,
                        JobPhoto.stage == photo_stage,
                        JobPhoto.uploaded_at >= changed_at_utc - timedelta(minutes=15),
                        JobPhoto.uploaded_at <= changed_at_utc + timedelta(minutes=15)
                    ).order_by(JobPhoto.uploaded_at)
                    
                    # Debug: Log the query parameters
                    logging.debug(f"Photo query for job {record.job_id}, stage {photo_stage}")
                    logging.debug(f"Changed at: {record.changed_at}")
                    logging.debug(f"Changed at UTC: {changed_at_utc}")
                    logging.debug(f"Time window: {changed_at_utc - timedelta(minutes=15)} to {changed_at_utc + timedelta(minutes=15)}")
                    
                    photos = photo_query.all()
                    
                    # Debug: Log the number of photos found
                    logging.debug(f"Found {len(photos)} photos")
                    
                    # Format photo data for attachments
                    for photo in photos:
                        filename = os.path.basename(photo.file_path)
                        file_url = url_for('uploaded_file', filename=filename, _external=True) if filename else None
                        attachments.append({
                            'id': photo.id,
                            'stage': photo.stage,
                            'file_url': file_url,
                            'uploaded_at': photo.uploaded_at.isoformat() if photo.uploaded_at else None
                        })
            
            audit_entry = {
                'id': record.id,
                'job_id': record.job_id,
                'changed_at': record.changed_at.isoformat() if record.changed_at else None,
                'changed_by': record.changed_by,
                'changed_by_name': changed_by_name,
                'changed_by_email': record.changed_by_user.email if record.changed_by_user else None,
                'role': role,
                'old_status': record.old_status,
                'new_status': record.new_status,
                'old_status_label': old_status_label,
                'new_status_label': new_status_label,
                'status_label': new_status_label,
                'reason': record.reason,
                'remark': remark,
                'description': generate_change_description(record),
                'attachments': attachments,
                'extra_services': extra_services if extra_services is not None else [],
                'driver_remarks': remarks_list if remarks_list else []
            }
            

            audit_data.append(audit_entry)
        
        # Prepare job driver information
        driver_info = None
        if job.driver:
            driver_info = {
                'id': job.driver.id,
                'name': job.driver.name,
                'license_number': job.driver.license_number
            }
        
        return jsonify({
            'audit_records': audit_data,
            'driver_info': driver_info
        }), 200
        
    except Exception as e:
        logging.error(f"Error retrieving job audit records: {e}", exc_info=True)
        return jsonify({'error': 'An error occurred while retrieving audit records'}), 500


@job_bp.route('/jobs/bulk-cancel', methods=['POST'])
@roles_accepted('admin', 'manager', 'accountant')
def bulk_cancel_jobs():
    """Cancel multiple jobs in bulk with a single reason."""
    try:
        # Get JSON data from request
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
            
        # Validate required fields
        job_ids = data.get('job_ids')
        reason = data.get('reason')
        
        if not job_ids or not isinstance(job_ids, list) or len(job_ids) == 0:
            return jsonify({'error': 'job_ids array is required and must contain at least one job ID'}), 400
            
        if not reason:
            return jsonify({'error': 'reason is required'}), 400
            
        # Validate that all job IDs are integers
        try:
            job_ids = [int(job_id) for job_id in job_ids]
        except (ValueError, TypeError):

            return jsonify({'error': 'All job_ids must be valid integers'}), 400
            
        # Get all jobs to be canceled
        jobs = Job.query.filter(Job.id.in_(job_ids), Job.is_deleted.is_(False)).all()
        
        # Check if all requested jobs were found
        found_job_ids = [job.id for job in jobs]
        missing_job_ids = [job_id for job_id in job_ids if job_id not in found_job_ids]
        
        if missing_job_ids:
            return jsonify({
                'error': f'Some jobs were not found: {missing_job_ids}'
            }), 404
            
        # Check if any jobs are already canceled
        already_canceled_jobs = [job for job in jobs if job.status == JobStatus.CANCELED.value]
        if already_canceled_jobs:
            already_canceled_ids = [job.id for job in already_canceled_jobs]
            return jsonify({
                'error': f'Some jobs are already canceled: {already_canceled_ids}'
            }), 400
            
        # Create audit records and update job statuses
        canceled_jobs = []
        audit_records = []
        audit_record_ids = []
        
        try:
            # Since Flask-SQLAlchemy auto-starts transactions, we work within that context
            # to ensure atomicity without creating nested transactions
            for job in jobs:
                # Store the previous status
                old_status = job.status
                
                # Update job status to canceled
                job.status = JobStatus.CANCELED.value
                
                # Create audit record
                audit_record = JobAudit(
                    job_id=job.id,
                    changed_by=current_user.id if current_user.is_authenticated else None,
                    old_status=old_status,
                    new_status=JobStatus.CANCELED.value,
                    reason=reason
                )
                
                audit_records.append(audit_record)
                canceled_jobs.append({
                    'job_id': job.id,
                    'old_status': old_status,
                    'new_status': JobStatus.CANCELED.value
                })
            
            # Add all audit records to database
            db.session.add_all(audit_records)
            
            # Clear monitoring alerts for canceled jobs
            from backend.models.job_monitoring_alert import JobMonitoringAlert
            JobMonitoringAlert.clear_alerts_for_canceled_jobs([job.id for job in jobs])
            
            # Commit to get the audit record IDs
            db.session.commit()
            
            # Get the IDs of the created audit records
            audit_record_ids = [record.id for record in audit_records]
        except Exception as e:
            # Explicitly rollback on error since we're in Flask-SQLAlchemy's auto-transaction
            db.session.rollback()
            raise e
        
        return jsonify({
            'message': f'Successfully canceled {len(canceled_jobs)} jobs',
            'canceled_jobs': canceled_jobs,
            'audit_record_ids': audit_record_ids
        }), 200
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error in bulk job cancellation: {str(e)}", exc_info=True)
        return jsonify({'error': f'An error occurred while canceling jobs: {str(e)}'}), 500


@job_bp.route('/jobs/reinstate/<int:job_id>', methods=['POST'])
@roles_accepted('admin', 'manager', 'accountant')
def reinstate_job(job_id):
    """Reinstate a canceled job to its previous status."""
    try:
        # Get the job to ensure it exists
        job = Job.query.get(job_id)
        if not job:
            return jsonify({'error': 'Job not found'}), 404
            
        # Check if job is actually canceled
        if job.status != JobStatus.CANCELED.value:
            return jsonify({
                'error': f'Job is not canceled (current status: {job.status})'
            }), 400
            
        # Find the most recent cancellation record for this job
        cancellation_audit = JobAudit.query.filter_by(
            job_id=job_id,
            new_status=JobStatus.CANCELED.value
        ).order_by(JobAudit.changed_at.desc()).first()
        
        # Handle legacy canceled jobs without audit records
        if not cancellation_audit:
            db.session.rollback()
            logging.error(f"No cancellation record found for job {job_id} - cannot reinstate without prior status")
            return jsonify({'error': 'Cannot reinstate: No cancellation record found'}), 400
        else:
            previous_status = cancellation_audit.old_status
        
        # Validate that the previous status is a valid job status
        valid_statuses = [status.value for status in JobStatus]
        if previous_status not in valid_statuses:
            db.session.rollback()
            return jsonify({
                'error': f'Invalid previous status: {previous_status}'
            }), 400
            
        # Validate context - check if job pickup date has passed
        if job.pickup_date:
            try:
                from datetime import datetime
                pickup_date = datetime.strptime(job.pickup_date, '%Y-%m-%d').date()
                today = datetime.now().date()
                if pickup_date < today and previous_status in ['new', 'pending', 'confirmed', 'otw']:
                    # Downgrade to 'new' status if pickup date has passed
                    previous_status = 'new'
            except ValueError:
                # If date parsing fails, log warning but continue with original status
                logging.warning(f"Could not parse pickup_date '{job.pickup_date}' for job {job_id}")
        
        # Store current status before changing it (for audit)
        old_status = job.status
        
        # Update job status to previous status (or downgraded status)
        job.status = previous_status
        
        # Create audit record for reinstatement
        audit_record = JobAudit(
            job_id=job.id,
            changed_by=current_user.id if current_user.is_authenticated else None,
            old_status=old_status,
            new_status=previous_status,
            reason='Job reinstated'
        )
        
        # Add audit record to database
        db.session.add(audit_record)
        db.session.commit()
        
        return jsonify({
            'message': f'Job successfully reinstated to {previous_status} status',
            'job_id': job.id,
            'previous_status': previous_status,
            'audit_record_id': audit_record.id
        }), 200
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error in job reinstatement: {e}", exc_info=True)
        return jsonify({'error': 'An error occurred while reinstating the job'}), 500

@job_bp.route('/jobs/calendar', methods=['GET'])
@auth_required()
def jobs_calendar():
    """
    Return jobs grouped by driver and date for the next *n* days (default 2) from a given start date.
    If no start_date is provided, defaults to today.

    Query parameters:
    - days: number of days to show (default 2, max 7)
    - start_date: YYYY-MM-DD format (optional, defaults to today)

    Response structure:
    {
      "calendar_data": {
        "YYYY-MM-DD": {
          "<driver_id>": [ {...job fields...} ],
          "unassigned":  [ {...job fields...} ]
        },
        ...
      },
      "drivers": [{"id": 1, "name": "John Doe"}, ...],
      "date_range": ["YYYY-MM-DD", "YYYY-MM-DD"]
    }
    """
    try:
        # ------------------------------------------------------------------
        # 1. Parse & clamp query params
        # ------------------------------------------------------------------
        days_param = request.args.get('days', 2)
        start_date_param = request.args.get('start_date')
        
        try:
            days_ahead = int(days_param)
        except ValueError:
            days_ahead = 2
        days_ahead = max(1, min(days_ahead, 7))  # prevent abuse

        # ------------------------------------------------------------------
        # 2. Build date range (ISO strings to match Job pickup_date)
        # ------------------------------------------------------------------
        if start_date_param:
            try:
                # Parse the provided start date
                start_date = datetime.strptime(start_date_param, '%Y-%m-%d').date()
            except ValueError:
                # Invalid date format, fallback to today
                start_date = datetime.utcnow().date()
        else:
            # Default to today if no start_date provided
            start_date = datetime.utcnow().date()
            
        date_range = [
            (start_date + timedelta(days=i)).strftime('%Y-%m-%d')
            for i in range(days_ahead)
        ]

        # ------------------------------------------------------------------
        # 3. Query jobs in range and allowed statuses
        # ------------------------------------------------------------------
        allowed_statuses = ['new', 'pending', 'confirmed', 'otw', 'ots', 'pob']
        jobs_query = (
            Job.query
            .filter(Job.pickup_date.in_(date_range))
            .filter(Job.status.in_(allowed_statuses))
            .options(
                db.joinedload(Job.driver),
                db.joinedload(Job.customer)
            )
        )
        jobs = jobs_query.all()

        # ------------------------------------------------------------------
        # 4. Build calendar_data & drivers lookup
        # ------------------------------------------------------------------
        calendar_data = {d: {} for d in date_range}
        drivers_lookup = {}

        for job in jobs:
            date_key = job.pickup_date
            # Use numeric driver_id directly; 0 for unassigned (more type-safe than string mixing)
            driver_key = job.driver_id if job.driver_id else 0

            # Init nested dicts
            calendar_data.setdefault(date_key, {})
            calendar_data[date_key].setdefault(driver_key, [])

            # Append job details required by frontend calendar
            calendar_data[date_key][driver_key].append({
                'id': job.id,
                'customer_name': job.customer.name if job.customer else None,
                'service_type': job.service_type,
                'pickup_time': job.pickup_time,
                'pickup_location': job.pickup_location,
                'dropoff_location': job.dropoff_location,
                'status': job.status
            })

            # Track driver meta once (skip unassigned which is 0)
            if driver_key != 0 and driver_key not in drivers_lookup:
                drivers_lookup[driver_key] = {
                    'id': job.driver_id,
                    'name': job.driver.name if job.driver else 'Unknown'
                }

        # ------------------------------------------------------------------
        # 5. Ensure *all* active drivers are returned (even if no jobs)
        # ------------------------------------------------------------------
        all_active_drivers = (
            Driver.query
            .filter_by(status='Active')
            .with_entities(Driver.id, Driver.name)
            .all()
        )
        drivers_full = [{'id': d.id, 'name': d.name} for d in all_active_drivers]

        return jsonify({
            'calendar_data': calendar_data,
            'drivers': drivers_full,
            'date_range': date_range
        }), 200

    except Exception as e:
        logging.error(f"Unhandled error in jobs_calendar: {e}", exc_info=True)
        return jsonify({'error': 'An unexpected error occurred. Please try again later.'}), 500


@job_bp.route('/jobs/revalidate', methods=['POST'])
@auth_required()
def revalidate_data():
    """Re-validate data with new column mappings"""
    try:
        data = request.get_json()
        column_mapping = data.get('column_mapping', {})
        row_data = data.get('data', [])
        
        # Re-validate data with new column mappings
        preview_data = revalidate_data_with_mapping(row_data, column_mapping)
        
        return jsonify({
            'success': True,
            'preview_data': preview_data
        })
        
    except Exception as e:
        logging.error(f"Error revalidating data: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'message': 'An error occurred while re-validating data. Please try again.'
        })


@job_bp.route('/jobs/validate-row', methods=['POST'])
@auth_required()
def validate_single_row():
    """Validate a single row against the database"""
    try:
        data = request.get_json()
        row_data = data.get('row_data', {})
        
        # Validate single row against database
        validation_result = validate_single_row_data(row_data)
        
        return jsonify({
            'success': True,
            'is_valid': validation_result['is_valid'],
            'error_message': validation_result['error_message']
        })
        
    except Exception as e:
        logging.error(f"Error validating single row: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'message': 'An error occurred while validating the row. Please try again.'
        })


def revalidate_data_with_mapping(row_data, column_mapping):
    """Re-validate data with new column mappings"""
    try:
        # Get lookup data using centralized validation
        lookups = get_validation_lookups()
        
        valid_count = 0
        error_count = 0
        validated_rows = []
        
        for row in row_data:
            # Apply column mapping if provided
            mapped_row = {
                'customer': row.get(column_mapping.get('customer', 'customer'), row.get('customer', '')),
                'customer_reference_no': row.get(column_mapping.get('customer_reference_no', 'customer_reference_no'), row.get('customer_reference_no', '')),
                'department': row.get(column_mapping.get('department', 'department'), row.get('department', '')),
                'service': row.get(column_mapping.get('service', 'service'), row.get('service', '')),
                'vehicle': row.get(column_mapping.get('vehicle', 'vehicle'), row.get('vehicle', '')),
                'driver': row.get(column_mapping.get('driver', 'driver'), row.get('driver', '')),
                'contractor': row.get(column_mapping.get('contractor', 'contractor'), row.get('contractor', '')),
                'vehicle_type': row.get(column_mapping.get('vehicle_type', 'vehicle_type'), row.get('vehicle_type', '')),
                'pickup_date': row.get('pickup_date', ''),
                'pickup_time': row.get('pickup_time', ''),
                'pickup_location': row.get('pickup_location', ''),
                'dropoff_location': row.get('dropoff_location', ''),
                'passenger_name': row.get('passenger_name', ''),
                'passenger_mobile': row.get(column_mapping.get('passenger_mobile', 'passenger_mobile'), row.get('passenger_mobile', '')),
                'status': row.get('status', 'new'),
                'remarks': row.get('remarks', ''),
                'row_number': row.get('row_number', 1),
            }

            # Use centralized validation
            is_valid, error_message, validated_data = validate_job_row(mapped_row, lookups)

            # Update row with validation results
            validated_row = {
                'row_number': mapped_row['row_number'],
                'customer': mapped_row['customer'],
                'customer_reference_no': mapped_row['customer_reference_no'],
                'department': mapped_row['department'],
                'service': mapped_row['service'],
                'vehicle': mapped_row['vehicle'],
                'driver': mapped_row['driver'],
                'contractor': mapped_row['contractor'],
                'vehicle_type': mapped_row['vehicle_type'],
                'pickup_date': mapped_row['pickup_date'],
                'pickup_time': mapped_row['pickup_time'],
                'pickup_location': mapped_row['pickup_location'],
                'dropoff_location': mapped_row['dropoff_location'],
                'passenger_name': mapped_row['passenger_name'],
                'passenger_mobile': mapped_row['passenger_mobile'],
                'status': mapped_row['status'],
                'remarks': mapped_row['remarks'],
                'is_valid': is_valid,
                'error_message': error_message
            }

            # Add validated IDs if available
            if 'customer_id' in validated_data:
                validated_row['customer_id'] = validated_data['customer_id']
            if 'service_id' in validated_data:
                validated_row['service_id'] = validated_data['service_id']
            if 'driver_id' in validated_data:
                validated_row['driver_id'] = validated_data['driver_id']
            if 'vehicle_id' in validated_data:
                validated_row['vehicle_id'] = validated_data['vehicle_id']
            if 'contractor_id' in validated_data:
                validated_row['contractor_id'] = validated_data['contractor_id']
            if 'vehicle_type_id' in validated_data:
                validated_row['vehicle_type_id'] = validated_data['vehicle_type_id']
            
            if is_valid:
                valid_count += 1
            else:
                error_count += 1
            
            validated_rows.append(validated_row)
        
        return {
            'valid_count': valid_count,
            'error_count': error_count,
            'rows': validated_rows,
            'json_data': json.dumps(validated_rows)
        }
        
    except Exception as e:
        logging.error(f"Error re-validating data: {str(e)}", exc_info=True)
        return {
            'valid_count': 0,
            'error_count': len(row_data),
            'rows': [],
            'json_data': '[]'
        }


def validate_single_row_data(row_data):
    """Validate a single row against the database"""
    try:
        # Get lookup data using centralized validation
        lookups = get_validation_lookups()
        
        # Use centralized validation
        is_valid, error_message, validated_data = validate_job_row(row_data, lookups)
        
        return {
            'is_valid': is_valid,
            'error_message': error_message
        }
        
    except Exception as e:
        logging.error(f"Error validating single row: {str(e)}", exc_info=True)
        return {
            'is_valid': False,
            'error_message': 'An error occurred during validation. Please try again.'
        }


@job_bp.route('/jobs/download-selected', methods=['POST'])
@auth_required()
def download_selected_rows():
    """Download selected rows as Excel file"""
    try:
        data = request.get_json()
        selected_rows = data.get('selected_rows', [])
        
        if not selected_rows:
            return jsonify({
                'success': False,
                'message': 'No rows selected for download'
            }), 400
        
        # Create a DataFrame from selected rows
        df = pd.DataFrame(selected_rows)
        
        # Create Excel file in memory
        output = BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name='Selected Jobs', index=False)
        
        output.seek(0)
        
        # Create response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=selected_jobs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        return response
        
    except Exception as e:
        logging.error(f"Error downloading selected rows: {str(e)}", exc_info=True)
        return jsonify({
            'success': False,
            'message': 'An error occurred while downloading the file. Please try again.'
        }), 500 
        return jsonify({'error': 'An unexpected error occurred. Please try again later.'}), 500 

@job_bp.route('/jobs/unbilled', methods=['GET'])
@roles_accepted('admin', 'manager', 'accountant')
def jobs_table_unbilled():
    try:
        page = int(request.args.get('page', 1))
        page_size = int(request.args.get('pageSize', 50))
        customer_id_param = request.args.get('customer_id')
        customer_id = None
        if customer_id_param:
            try:
                customer_id = int(customer_id_param)
            except ValueError:
                return jsonify({'error': 'Invalid customer_id'}), 400
        query = Job.query.filter(
            Job.invoice_id == None,
            Job.status == 'jc'
        )
        if customer_id is not None:
            query = query.filter(Job.customer_id == customer_id)
        total = query.count()
        jobs = (
            query.order_by(Job.pickup_date.desc(), Job.pickup_time.desc())
            .offset((page - 1) * page_size)
            .limit(page_size)
            .all()
        )
        response_json = {
            'items': schema_many.dump(jobs),
            'total': total,
            'page': page,
            'pageSize': page_size
        }
        return jsonify(response_json), 200

    except Exception as e:
        logging.error(f"Unhandled error in jobs_table_unbilled: {e}", exc_info=True)
        return jsonify({'error': 'An unexpected error occurred. Please try again later.'}), 500

@job_bp.route('/jobs/contractor-billable', methods=['GET'])
@roles_accepted('admin', 'manager', 'accountant')
def jobs_contractor_billable():
    try:
        contractor_id = request.args.get('contractor_id', type=int)
        
        # Get billable jobs from BillService
        billable_jobs = BillService.get_billable_jobs(contractor_id)
        
        response_json = {
            'items': schema_many.dump(billable_jobs),
            'total': len(billable_jobs)
        }
        return jsonify(response_json), 200

    except ServiceError as se:
        logging.error(f"ServiceError in jobs_contractor_billable: {se.message}")
        return jsonify({'error': se.message}), 400
    except Exception as e:
        logging.error(f"Unhandled error in jobs_contractor_billable: {e}", exc_info=True)
        return jsonify({'error': 'An unexpected error occurred. Please try again later.'}), 500

@job_bp.route('/jobs/driver-billable', methods=['GET'])
@roles_accepted('admin', 'manager', 'accountant')
def jobs_driver_billable():
    try:
        driver_id = request.args.get('driver_id', type=int)
        
        # Get billable jobs from BillService
        billable_jobs = BillService.get_driver_billable_jobs(driver_id)
        
        response_json = {
            'items': schema_many.dump(billable_jobs),
            'total': len(billable_jobs)
        }
        return jsonify(response_json), 200

    except ServiceError as se:
        logging.error(f"ServiceError in jobs_driver_billable: {se.message}")
        return jsonify({'error': se.message}), 400
    except Exception as e:
        logging.error(f"Unhandled error in jobs_driver_billable: {e}", exc_info=True)
        return jsonify({'error': 'An unexpected error occurred. Please try again later.'}), 500

@job_bp.route('/jobs/remove/<int:id>', methods=['DELETE'])
@roles_accepted('admin', 'manager', 'accountant')
def remove_job_from_invoice(id):
    try:
        report = JobService.remove_job_from_invoice(id)
        return jsonify(report), 200
    except ServiceError as se:
        return jsonify({'error': se.message}), 400
    except Exception as e:
        logging.error(f"Unhandled error in billing_report: {e}", exc_info=True)
        return jsonify({'error': 'An unexpected error occurred. Please try again later.'}), 500 

@job_bp.route('/jobs/update/<int:job_id>', methods=['PUT'])
@roles_accepted('admin', 'manager', 'accountant')
def update_job_and_invoice(job_id):
    try:
        data = request.get_json()
        errors = schema.validate(data, partial=True)
        if errors:
            return jsonify(errors), 400
        result = JobService.updateJobAndInvoice(job_id, data)
        if isinstance(result, dict) and result.get('error'):
            return jsonify(result), 400
        if not result:
            return jsonify({'error': 'Job not found'}), 404
        return jsonify(schema.dump(result)), 200
    except ServiceError as se:
        return jsonify({'error': se.message}), 400
    except Exception as e:
        logging.error(f"Unhandled error in update_job: {e}", exc_info=True)
        return jsonify({'error': 'An unexpected error occurred. Please try again later.'}), 500



# ---------------- PIN CODE LOOKUP ----------------
@limiter.limit("10 per minute")
@job_bp.route('/lookup/pincode', methods=['GET'])
@auth_required()
def lookup_pincode():
    """
    Pin Code Lookup API
    Accepts: postalcode (or zipcode), countrycode
    First checks local database, then calls OneMap API for Singapore or OpenStreetMap Nominatim for other countries

    This implementation addresses multiple critical issues:
    1. Robust error handling to prevent application crashes
    2. Input validation to prevent malformed data from reaching external APIs
    3. Efficient database lookups using indexed queries (O(1) performance)
    4. Graceful degradation when postal code data is unavailable
    5. OneMap API integration for Singapore postal codes
    """
    try:
        from backend.models.postal_code import PostalCode
        from backend.services.onemap_service import get_address_from_postal_code
    except ImportError as e:
        logging.error(f"Failed to import required modules: {e}")
        return jsonify({'error': 'Postal code service temporarily unavailable'}), 503

    # Extract and validate input parameters
    postalcode = request.args.get('postalcode') or request.args.get('zipcode')
    countrycode = request.args.get('countrycode')

    if not postalcode or not countrycode:
        return jsonify({'error': 'postalcode and countrycode are required'}), 400

    # Critical: Validate postal code format to prevent malformed input from reaching external APIs
    if not re.match(r'^[A-Za-z0-9\s\-]{1,20}$', postalcode):
        return jsonify({'error': 'Invalid postal code format'}), 400

    # Clean and normalize postal code for database lookup
    clean_postalcode = postalcode.strip().replace(' ', '').replace('-', '')

    try:
        # First, check local database for postal code (O(1) indexed lookup)
        logging.info(f"Checking local database for postal code: {clean_postalcode}")

        try:
            # Use indexed query for O(1) performance - postal_code column is indexed
            local_postal_code = PostalCode.query.filter_by(postal_code=clean_postalcode).first()
        except Exception as db_error:
            # Database lookup failed - log but don't crash, fall back to external API
            logging.warning(f"Database lookup failed for postal code {clean_postalcode}: {db_error}")
            local_postal_code = None

        if local_postal_code:
            logging.info(f"Found postal code in local database: {local_postal_code.postal_code}")

            # Check for potential data quality issues (duplicate postal codes)
            try:
                duplicate_count = PostalCode.query.filter_by(postal_code=clean_postalcode).count()
                if duplicate_count > 1:
                    logging.warning(
                        f"Duplicate postal code found in database: '{clean_postalcode}'. "
                        f"Found {duplicate_count} entries. Using the first one."
                    )
            except Exception as dup_check_error:
                logging.warning(f"Failed to check for duplicates: {dup_check_error}")

            # Return address from local database
            # For Singapore postal codes, parse the address to extract components
            address_parts = local_postal_code.address.split(',') if local_postal_code.address else []
            normalized = {
                'city': 'Singapore' if countrycode.lower() == 'sg' else None,
                'state': 'Singapore' if countrycode.lower() == 'sg' else None,
                'country': 'Singapore' if countrycode.lower() == 'sg' else None,
                'postcode': local_postal_code.postal_code,
                'locality': address_parts[0].strip() if address_parts else None,
                'display_name': local_postal_code.address,
                'lat': None,  # Local database doesn't store coordinates
                'lon': None,
                'source': 'local_database'
            }
            return jsonify({'address': normalized}), 200

        # If not found in local database, fall back to external API
        logging.info(f"[POSTAL CODE LOOKUP] Not found in local database, trying external APIs for: {postalcode} (country: {countrycode})")

        # For Singapore, try OneMap API first, then fall back to OpenStreetMap
        if countrycode.lower() == 'sg':
            try:
                logging.info(f"[ONEMAP API] Starting lookup for Singapore postal code: {clean_postalcode}")
                address = get_address_from_postal_code(clean_postalcode, 'Singapore')

                if address:
                    logging.info(f"[ONEMAP API] SUCCESS - Found address: {address}")
                    normalized = {
                        'city': 'Singapore',
                        'state': 'Singapore',
                        'country': 'Singapore',
                        'postcode': clean_postalcode,
                        'locality': None,
                        'display_name': address,
                        'lat': None,
                        'lon': None,
                        'source': 'onemap_api'
                    }

                    # Cache the result to local database
                    try:
                        existing = PostalCode.query.filter_by(postal_code=clean_postalcode).first()
                        if not existing:
                            new_postal_code = PostalCode(
                                postal_code=clean_postalcode,
                                address=address
                            )
                            db.session.add(new_postal_code)
                            db.session.commit()
                            logging.info(f"[ONEMAP API] Cached result to local database for postal code: {clean_postalcode}")
                    except Exception as cache_error:
                        logging.warning(f"[ONEMAP API] Failed to cache to database: {cache_error}")
                        try:
                            db.session.rollback()
                        except Exception:
                            pass

                    return jsonify({'address': normalized}), 200
                else:
                    # OneMap returned None, fall back to OpenStreetMap
                    logging.warning(f"[ONEMAP API] No results found for postal code: {clean_postalcode}")
                    logging.info(f"[OPENSTREETMAP] Falling back to OpenStreetMap for postal code: {clean_postalcode}")

            except Exception as onemap_error:
                # OneMap API failed, fall back to OpenStreetMap
                logging.error(f"[ONEMAP API] Error occurred: {onemap_error}")
                logging.info(f"[OPENSTREETMAP] Falling back to OpenStreetMap due to OneMap error")

        # Use OpenStreetMap Nominatim (for non-Singapore or as fallback for Singapore)
        try:
            logging.info(f"[OPENSTREETMAP] Starting lookup for postal code: {postalcode} (country: {countrycode})")
            url = "https://nominatim.openstreetmap.org/search"
            params = {
                'q': postalcode,  # Use original postal code for external API (may handle formatting)
                'countrycodes': countrycode,
                'format': 'json',
                'addressdetails': 1,
                'limit': 1
            }

            response = requests.get(
                url,
                params=params,
                headers={"User-Agent": "fleetwise/1.0"},
                timeout=10
            )
            response.raise_for_status()
            data = response.json()

            if not data:
                logging.warning(f"[OPENSTREETMAP] No results found for postal code: {postalcode}")
                return jsonify({'error': 'No address found for given postal code'}), 404

            logging.info(f"[OPENSTREETMAP] SUCCESS - Found address")
            address = data[0].get('address', {})
            normalized = {
                'city': address.get('city') or address.get('town') or address.get('village'),
                'state': address.get('state'),
                'country': address.get('country'),
                'postcode': address.get('postcode'),
                'locality': address.get('suburb') or address.get('locality'),
                'display_name': data[0].get('display_name'),
                'lat': data[0].get('lat'),
                'lon': data[0].get('lon'),
                'source': 'openstreetmap_api'
            }

            # Optionally cache the API result back to database for future lookups
            if normalized.get('display_name') and normalized.get('postcode'):
                try:
                    # Check if this postal code already exists (might have been added by another request)
                    existing = PostalCode.query.filter_by(postal_code=normalized['postcode']).first()
                    if not existing:
                        new_postal_code = PostalCode(
                            postal_code=normalized['postcode'],
                            address=normalized['display_name']
                        )
                        db.session.add(new_postal_code)
                        db.session.commit()
                        logging.info(f"[OPENSTREETMAP] Cached result to local database for postal code: {normalized['postcode']}")
                except Exception as cache_error:
                    logging.warning(f"[OPENSTREETMAP] Failed to cache to database: {cache_error}")
                    try:
                        db.session.rollback()
                    except Exception:
                        pass  # Ignore rollback errors
                    # Don't fail the request if caching fails

            return jsonify({'address': normalized}), 200

        except requests.Timeout:
            logging.error(f"[OPENSTREETMAP] Timeout for postal code: {postalcode}")
            return jsonify({'error': 'Address lookup service timeout - please try again'}), 504
        except requests.ConnectionError:
            logging.error(f"[OPENSTREETMAP] Connection error for postal code: {postalcode}")
            return jsonify({'error': 'Address lookup service unavailable - please check your connection'}), 503
        except requests.HTTPError as http_error:
            logging.error(f"[OPENSTREETMAP] HTTP error for postal code {postalcode}: {http_error}")
            return jsonify({'error': 'Address lookup service error - please try again later'}), 503
        except requests.RequestException as req_error:
            logging.error(f"[OPENSTREETMAP] Request error for postal code {postalcode}: {req_error}")
            return jsonify({'error': 'Address lookup service unavailable'}), 503

    except Exception as e:
        # Catch-all for any unexpected errors to prevent application crashes
        logging.error(f"Unexpected error in postal code lookup: {str(e)}", exc_info=True)
        return jsonify({
            'error': 'Internal server error during address lookup. Please try again later.'
        }), 500

@job_bp.route('/jobs/audit-trail', methods=['GET'])
@roles_accepted('admin', 'manager', 'driver', 'customer', 'accountant')
def get_jobs_audit_trail():
    """Get a summary list of jobs with audit changes, with filtering capabilities."""
    try:
        # ---------------------------------------------------------------------
        # Query Parameters
        # ---------------------------------------------------------------------
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        search = request.args.get('search', '').strip()

        try:
            page = int(request.args.get('page', 1))
            page_size = int(request.args.get('page_size', 50))
        except ValueError:
            return jsonify({'error': 'Invalid pagination parameters'}), 400

        page = max(1, page)
        page_size = max(1, min(page_size, 100))
        if page > 10000:
            return jsonify({'error': 'Page number exceeds maximum allowed value (10000)'}), 400

        # ---------------------------------------------------------------------
        # Step 1: Base Query — join Jobs with JobAudit
        # ---------------------------------------------------------------------
        query = (
            db.session.query(Job, JobAudit)
            .join(JobAudit, Job.id == JobAudit.job_id)
            .filter(Job.is_deleted.is_(False))
        )

        # ---------------------------------------------------------------------
        # Step 2: Role-Based Filtering (enforce RBAC directly here)
        # ---------------------------------------------------------------------
        if current_user.has_role('driver'):
            driver_id = getattr(current_user, 'driver_id', None)
            if not driver_id:
                abort(403, description='Driver profile missing')
            query = query.filter(Job.driver_id == driver_id)

        elif current_user.has_role('customer'):
            customer_id = getattr(current_user, 'customer_id', None)
            if not customer_id:
                abort(403, description='Customer profile missing')
            query = query.filter(Job.customer_id == customer_id)

        # ---------------------------------------------------------------------
        # Step 3: Date Range Filters
        # ---------------------------------------------------------------------
        if start_date:
            try:
                start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
                query = query.filter(JobAudit.changed_at >= start_date_obj)
            except ValueError:
                return jsonify({'error': 'Invalid start_date format. Use YYYY-MM-DD'}), 400

        if end_date:
            try:
                end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')
                end_date_obj = end_date_obj.replace(
                    hour=23, minute=59, second=59, microsecond=999999
                )
                query = query.filter(JobAudit.changed_at <= end_date_obj)
            except ValueError:
                return jsonify({'error': 'Invalid end_date format. Use YYYY-MM-DD'}), 400

        # ---------------------------------------------------------------------
        # Step 4: Search Filtering (job_id / customer / user)
        # ---------------------------------------------------------------------
        job_id_search = None
        if search:
            if search.startswith('JB-'):
                try:
                    job_id_search = int(search.split('-')[-1])
                except (ValueError, IndexError):
                    pass
            elif search.isdigit():
                job_id_search = int(search)

            search_conditions = []
            if job_id_search:
                search_conditions.append(Job.id == job_id_search)
            search_conditions.append(Customer.name.ilike(f'%{search}%'))
            search_conditions.append(User.email.ilike(f'%{search}%'))

            query = query.join(Customer, Job.customer_id == Customer.id)
            query = query.join(User, JobAudit.changed_by == User.id)
            query = query.filter(or_(*search_conditions))
        else:
            # Always include joins for output mapping
            query = query.join(Customer, Job.customer_id == Customer.id)
            query = query.join(User, JobAudit.changed_by == User.id)

        # ---------------------------------------------------------------------
        # Step 5: Latest Audit Record Subquery
        # ---------------------------------------------------------------------
        latest_audit_subquery = (
            db.session.query(
                JobAudit.job_id,
                func.max(JobAudit.changed_at).label('latest_change')
            )
            .group_by(JobAudit.job_id)
            .subquery()
        )

        query = query.join(
            latest_audit_subquery,
            and_(
                Job.id == latest_audit_subquery.c.job_id,
                JobAudit.changed_at == latest_audit_subquery.c.latest_change
            )
        ).order_by(JobAudit.changed_at.desc())

        # ---------------------------------------------------------------------
        # Step 6: Count & Pagination
        # ---------------------------------------------------------------------
        total_query = query.with_entities(func.count(Job.id.distinct()))
        total = total_query.scalar() or 0

        jobs_with_audit = (
            query
            .with_entities(Job, JobAudit, Customer, User)
            .offset((page - 1) * page_size)
            .limit(page_size)
            .all()
        )

        # ---------------------------------------------------------------------
        # Step 7: Format Output
        # ---------------------------------------------------------------------
        audit_summary = []
        for job, audit_record, customer, user in jobs_with_audit:
            audit_summary.append({
                'id': job.id,
                'job_id': job.id,
                'customer_name': customer.name if customer else None,
                'last_modified_date': (
                    audit_record.changed_at.isoformat()
                    if audit_record and audit_record.changed_at else None
                ),
                'last_change_made': generate_change_description(audit_record),
                'changed_by': {
                    'id': getattr(user, 'id', None),
                    'name': getattr(user, 'name', None) or getattr(user, 'email', None),
                    'email': getattr(user, 'email', None)
                }
            })

        # ---------------------------------------------------------------------
        # Step 8: Return JSON Response
        # ---------------------------------------------------------------------
        return jsonify({
            'items': audit_summary,
            'total': total,
            'page': page,
            'page_size': page_size
        }), 200

    except Exception as e:
        logging.error(f"Error retrieving jobs audit trail: {e}", exc_info=True)
        return jsonify({'error': 'An error occurred while retrieving audit trail'}), 500

def generate_change_description(audit_record):
    """Generate a human-readable description of the change from an audit record."""
    try:
        # Handle status changes
        if audit_record.old_status and audit_record.new_status:
            if audit_record.new_status == 'canceled' and audit_record.reason:
                return f"Canceled: {audit_record.reason}"
            return f"Status changed from {audit_record.old_status} to {audit_record.new_status}"
        
        # Handle cancellations
        if audit_record.new_status == 'canceled':
            if audit_record.reason:
                return f"Canceled: {audit_record.reason}"
            return "Job canceled"
            
        # Handle reinstatements
        if audit_record.reason and 'reinstated' in audit_record.reason.lower():
            return f"Job reinstated to {audit_record.new_status} status"
        
        # Handle penalty changes (check additional_data)
        if audit_record.additional_data:
            if isinstance(audit_record.additional_data, dict):
                if 'penalty' in audit_record.additional_data:
                    penalty = audit_record.additional_data['penalty']
                    return f"Penalty of ${penalty} added"
        
        # Default description
        if audit_record.reason:
            return audit_record.reason
            
        return "Job modified"
        
    except Exception as e:
        logging.error(f"Error generating change description: {e}")
        return "Job modified"


@job_bp.route('/jobs/update_status/<int:job_id>', methods=['POST'])
@roles_accepted('admin', 'manager')
def update_job_status(job_id):
    """
    Manually update a job's status with full audit trail.
    
    This endpoint allows admins/managers to manually advance a job's status
    when drivers cannot update via mobile app. Every change is recorded in
    the job_audit table for accountability.
    
    Request Body:
    {
        "new_status": "otw",           # Required - the new status to set
        "remark": "Driver confirmed via phone call"  # Optional - reason for manual update
    }
    
    Returns:
    {
        "message": "Job status updated successfully",
        "job_id": 123,
        "old_status": "confirmed",
        "new_status": "otw"
    }
    """
    try:
        # Get JSON data from request
        data = request.get_json()
        if not data:
            return jsonify({'error': 'No data provided'}), 400
            
        # Validate required fields
        new_status = data.get('new_status')
        if not new_status:
            return jsonify({'error': 'new_status is required'}), 400
            
        # Validate that the new status is a valid job status BEFORE acquiring lock
        # This prevents DoS attacks where invalid requests hold locks unnecessarily
        valid_statuses = [status.value for status in JobStatus]
        if new_status not in valid_statuses:
            return jsonify({
                'error': f'Invalid status. Must be one of: {", ".join(valid_statuses)}'
            }), 400
            
        # Get optional remark
        remark = data.get('remark')
        
        # Get optional changed_at timestamp from frontend
        changed_at_timestamp = data.get('changed_at')
        
        # Get the job with row-level locking to prevent race conditions
        # This is now done AFTER validation to minimize lock time
        job = Job.query.with_for_update().get(job_id)
        if not job:
            return jsonify({'error': 'Job not found'}), 404
            
        # Store the current job status before updating it
        old_status = job.status
        
        # Check if the status transition is valid according to business rules
        if not job.can_transition_to(new_status):
            return jsonify({
                'error': f'Invalid status transition from {old_status} to {new_status}'
            }), 400
            
        # All validations passed - proceed with update and audit record creation
        
        # Create audit record
        audit_record = JobAudit(
            job_id=job_id,
            changed_by=current_user.id if current_user.is_authenticated else None,
            old_status=old_status,
            new_status=new_status,
            reason=remark
        )
        
        # Set custom timestamp if provided by frontend
        if changed_at_timestamp:
            try:
                from datetime import datetime, timezone
                if isinstance(changed_at_timestamp, str):
                    # Try to parse as ISO format
                    try:
                        custom_timestamp = datetime.fromisoformat(changed_at_timestamp)
                    except ValueError:
                        # Try with Z suffix if original parsing fails
                        custom_timestamp = datetime.fromisoformat(changed_at_timestamp.replace('Z', '+00:00'))
                else:
                    # Assume it's already a datetime object
                    custom_timestamp = changed_at_timestamp
                
                # Handle timezone-naive timestamps by assuming local timezone
                if custom_timestamp.tzinfo is None:
                    # For naive timestamps, we'll do a more lenient comparison
                    # Get current time in local timezone
                    now_local = datetime.now()
                    # If the timestamp is reasonably close to now (within 24 hours), allow it
                    # This handles cases where frontend sends local time without timezone info
                    time_diff = now_local - custom_timestamp
                    if time_diff.total_seconds() < -86400:  # More than 24 hours in the future
                        return jsonify({
                            'success': False,
                            'message': 'Cannot set future timestamp for status change'
                        }), 400
                    # If it's not clearly in the future, treat it as valid
                else:
                    # For timezone-aware timestamps, do strict comparison
                    now = datetime.now(timezone.utc)
                    if custom_timestamp > now:
                        return jsonify({
                            'success': False,
                            'message': 'Cannot set future timestamp for status change'
                        }), 400
                
                # Validate chronological order with existing audit records
                last_audit = JobAudit.query.filter_by(job_id=job_id)\
                    .order_by(JobAudit.changed_at.desc())\
                    .first()
                if last_audit and last_audit.changed_at:
                    # Ensure last_audit.changed_at is timezone-aware
                    last_audit_timestamp = last_audit.changed_at
                    if last_audit_timestamp.tzinfo is None:
                        # If database timestamp is naive, assume it's UTC
                        last_audit_timestamp = last_audit_timestamp.replace(tzinfo=timezone.utc)
                    
                    # Ensure both timestamps are using the same timezone-aware format for comparison
                    if custom_timestamp.tzinfo is None:
                        # This shouldn't happen with our earlier conversion, but just in case
                        custom_timestamp = custom_timestamp.replace(tzinfo=timezone.utc)
                    
                    # Convert both to UTC for comparison if they have different timezones
                    if custom_timestamp.tzinfo != last_audit_timestamp.tzinfo:
                        custom_timestamp = custom_timestamp.astimezone(timezone.utc)
                        last_audit_timestamp = last_audit_timestamp.astimezone(timezone.utc)
                    
                    if custom_timestamp <= last_audit_timestamp:
                        return jsonify({
                            'success': False,
                            'message': f'Timestamp must be after last status change at {last_audit.changed_at.isoformat()}'
                        }), 400
                
                # Set the changed_at field directly
                audit_record.changed_at = custom_timestamp
            except ValueError as e:
                logging.warning(f"Invalid changed_at timestamp format: {changed_at_timestamp}, error: {e}")
                return jsonify({
                    'success': False,
                    'message': 'Invalid timestamp format. Use ISO 8601 format (YYYY-MM-DDTHH:MM:SS)'
                }), 400
        
        # Update job status
        job.status = new_status
        
        # Clear monitoring alerts if job status is updated to OTW
        if new_status == JobStatus.OTW.value:
            from backend.models.job_monitoring_alert import JobMonitoringAlert
            JobMonitoringAlert.clear_alert(job_id)
        
        # Add audit record to database
        db.session.add(audit_record)
        
        # Commit both changes atomically
        db.session.commit()
        
        # Send push notification to driver if assigned AFTER the transaction is committed
        # This ensures that even if notification fails, the job status update is preserved
        if job.driver_id:
            try:
                driver = Driver.query.get(job.driver_id)
                if driver:
                    user = User.query.filter_by(driver_id=driver.id).first()
                    if user:
                        # Import PushNotificationService here to avoid circular imports
                        from backend.services.push_notification_service import PushNotificationService
                        for token in [user.android_device_token, user.ios_device_token]:
                            if token:
                                notification_success = PushNotificationService.send(
                                    token=token,
                                    title="Job Status Updated",
                                    body=f"Job #{job.id} status updated to {new_status.title()}",
                                    data={"job_id": str(job.id), "status": new_status}
                                )
                                if not notification_success:
                                    logging.warning(f"Failed to send push notification for job {job.id}")
            except Exception as e:
                # Log the error but don't fail the request since the job status was successfully updated
                logging.warning(f"Exception while sending push notification for job {job.id}: {e}")
        
        return jsonify({
            'message': 'Job status updated successfully',
            'job_id': job_id,
            'old_status': old_status,
            'new_status': new_status
        }), 200
        
    except Exception as e:
        db.session.rollback()
        logging.error(f"Error updating job status: {e}", exc_info=True)
        return jsonify({'error': 'An error occurred while updating job status'}), 500
