"""
Script to generate 100 test jobs in Excel format for bulk upload functionality testing.

This script creates an Excel file with 100 sample job entries using actual database data
from customers, services, vehicles, drivers, contractors, and vehicle types.

Usage:
    # Make sure to activate the virtual environment first
    # On Windows:
    venv\\Scripts\\activate
    # On Unix/MacOS:
    source venv/bin/activate

    # Then run:
    python scripts/generate_bulk_upload_test_data.py

Output:
    Creates 'bulk_upload_100_jobs_test.xlsx' in the current directory
"""

import sys
import os

# Add parent directory to path to import backend modules
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), '..')))

from datetime import datetime, timedelta
import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from io import BytesIO

try:
    from backend.extensions import db
    from backend.models.customer import Customer
    from backend.models.driver import Driver
    from backend.models.service import Service
    from backend.models.vehicle import Vehicle
    from backend.models.contractor import Contractor
    from backend.models.vehicle_type import VehicleType
    from flask import Flask
except ImportError as e:
    print(f"Error importing modules: {e}")
    print("\nMake sure you have activated the virtual environment:")
    print("  On Windows: venv\\Scripts\\activate")
    print("  On Unix/MacOS: source venv/bin/activate")
    sys.exit(1)


def create_app():
    """Create Flask app instance for database access"""
    from dotenv import load_dotenv

    # Load environment variables
    load_dotenv()

    app = Flask(__name__)

    # Load configuration - use DevConfig which has SQLALCHEMY_DATABASE_URI set
    from backend.config import DevConfig
    app.config.from_object(DevConfig)

    # Initialize extensions
    db.init_app(app)

    return app


def generate_test_jobs(num_jobs=100):
    """
    Generate test job data for bulk upload testing

    Args:
        num_jobs (int): Number of test jobs to generate (default: 100)

    Returns:
        pd.DataFrame: DataFrame containing test job data
    """
    app = create_app()

    with app.app_context():
        # Fetch active records from database
        customers = Customer.query.filter_by(status='Active').all()
        services = Service.query.filter_by(status='Active').all()
        vehicles = Vehicle.query.filter_by(status='Active').all()
        drivers = Driver.query.filter_by(status='Active').all()
        contractors = Contractor.query.filter_by(status='Active').all()
        vehicle_types = VehicleType.query.filter_by(status=True, is_deleted=False).all()

        # Validate we have enough data
        if not all([customers, services, vehicles, drivers]):
            raise ValueError("Insufficient data in database. Need at least 1 active customer, service, vehicle, and driver.")

        print(f"Found {len(customers)} customers, {len(services)} services, {len(vehicles)} vehicles, {len(drivers)} drivers")
        print(f"Found {len(contractors)} contractors, {len(vehicle_types)} vehicle types")

        # Prepare data for cycling through
        customer_list = [c.name for c in customers]
        service_list = [s.name for s in services]
        vehicle_list = [v.number for v in vehicles]
        driver_list = [d.name for d in drivers]
        contractor_list = [c.name for c in contractors] if contractors else ['']
        vehicle_type_list = [vt.name for vt in vehicle_types] if vehicle_types else ['']

        # Sample locations for variety
        pickup_locations = [
            'Changi Airport Terminal 1',
            'Changi Airport Terminal 2',
            'Changi Airport Terminal 3',
            'Marina Bay Sands',
            'Raffles Place',
            'Orchard Road',
            'Jurong East',
            'Woodlands',
            'Tampines',
            'Bedok',
            'Sembawang',
            'Pasir Ris',
            'Clementi',
            'Bishan',
            'Ang Mo Kio',
            'Toa Payoh',
            'Novena',
            'Dhoby Ghaut',
            'City Hall',
            'Bukit Batok'
        ]

        dropoff_locations = [
            'Sentosa Island',
            'Gardens by the Bay',
            'Universal Studios Singapore',
            'Clarke Quay',
            'Chinatown',
            'Little India',
            'Bugis',
            'Suntec City',
            'VivoCity',
            'Harbourfront',
            'NUS',
            'NTU',
            'Singapore Zoo',
            'Night Safari',
            'East Coast Park',
            'Punggol Waterway',
            'Jewel Changi',
            'ION Orchard',
            'Marina Square',
            'Esplanade'
        ]

        departments = [
            'Operations Department',
            'Sales Department',
            'IT Department',
            'Marketing Department',
            'Finance Department',
            'HR Department',
            'Logistics Department',
            'Customer Service',
            'Administration',
            'Procurement'
        ]

        # Generate test data
        import random

        test_jobs = []
        today = datetime.now()

        # Define error scenarios first
        error_scenarios = [
            {'Customer': 'INVALID_CUSTOMER', 'error': 'Invalid customer name'},
            {'Service': 'INVALID_SERVICE', 'error': 'Invalid service name'},
            {'Vehicle': 'INVALID_VEHICLE', 'error': 'Invalid vehicle number'},
            {'Driver': 'INVALID_DRIVER', 'error': 'Invalid driver name'},
            {'Pickup Date': '2025-13-45', 'error': 'Invalid date format'},
            {'Pickup Time': '25:99', 'error': 'Invalid time format'},
            {'Passenger Mobile': 'INVALID_PHONE', 'error': 'Invalid phone number'},
            {'Customer Reference No': '', 'error': 'Missing reference number'},
            {'Pickup Location': '', 'error': 'Missing pickup location'},
            {'Drop-off Location': '', 'error': 'Missing dropoff location'}
        ]

        # Generate random positions for invalid jobs (0-99)
        # Ensure they're spread throughout the dataset, not clustered
        random.seed(42)  # Set seed for reproducibility
        invalid_positions = sorted(random.sample(range(100), 10))

        print(f"Invalid jobs will be placed at positions: {[p+1 for p in invalid_positions]}")

        # Generate all 100 jobs
        error_idx = 0
        for i in range(100):
            pickup_date = today + timedelta(days=(i % 30) + 1)
            hour = 8 + (i % 12)  # Hours between 8 AM and 8 PM
            minute = (i % 4) * 15  # 0, 15, 30, 45 minutes

            # Check if this position should have an invalid job
            if i in invalid_positions:
                # Create invalid job
                error_scenario = error_scenarios[error_idx]
                job = {
                    'Customer': customer_list[0],
                    'Customer Reference No': f'REF{str(i+1).zfill(4)}',
                    'Department/Person In Charge/Sub-Customer': 'Testing Department',
                    'Service': service_list[0],
                    'Vehicle': vehicle_list[0],
                    'Driver': driver_list[0],
                    'Contractor': contractor_list[0] if contractor_list else '',
                    'Vehicle Type': vehicle_type_list[0] if vehicle_type_list else '',
                    'Pickup Date': pickup_date.strftime('%Y-%m-%d'),
                    'Pickup Time': '09:00',
                    'Pickup Location': f'Test Pickup Location {i+1}',
                    'Drop-off Location': f'Test Dropoff Location {i+1}',
                    'Passenger Name': f'Test Passenger {i+1}',
                    'Passenger Mobile': f'+6590000{str(error_idx).zfill(3)}',
                    'Remarks': f'INVALID DATA TEST - {error_scenario["error"]} - Row {i+1}'
                }

                # Apply the error scenario
                job.update(error_scenario)
                if 'error' in job:
                    del job['error']

                error_idx += 1
            else:
                # Create valid job
                job = {
                    'Customer': customer_list[i % len(customer_list)],
                    'Customer Reference No': f'REF{str(i+1).zfill(4)}',
                    'Department/Person In Charge/Sub-Customer': departments[i % len(departments)],
                    'Service': service_list[i % len(service_list)],
                    'Vehicle': vehicle_list[i % len(vehicle_list)],
                    'Driver': driver_list[i % len(driver_list)],
                    'Contractor': contractor_list[i % len(contractor_list)],
                    'Vehicle Type': vehicle_type_list[i % len(vehicle_type_list)],
                    'Pickup Date': pickup_date.strftime('%Y-%m-%d'),
                    'Pickup Time': f'{str(hour).zfill(2)}:{str(minute).zfill(2)}',
                    'Pickup Location': pickup_locations[i % len(pickup_locations)],
                    'Drop-off Location': dropoff_locations[i % len(dropoff_locations)],
                    'Passenger Name': f'Passenger {i+1}',
                    'Passenger Mobile': f'+659{str(1000000 + i).zfill(7)}',
                    'Remarks': f'Test job entry {i+1} - Valid data for bulk upload testing'
                }

            test_jobs.append(job)

        return pd.DataFrame(test_jobs), {
            'customers': customer_list,
            'services': service_list,
            'vehicles': vehicle_list,
            'drivers': driver_list,
            'contractors': contractor_list,
            'vehicle_types': vehicle_type_list
        }


def create_excel_with_formatting(df, validation_data, output_filename='bulk_upload_100_jobs_test.xlsx'):
    """
    Create Excel file with proper formatting and data validation

    Args:
        df (pd.DataFrame): DataFrame with job data
        validation_data (dict): Dictionary containing validation lists
        output_filename (str): Output filename
    """
    with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Jobs Template', index=False)

        # Get the workbook and worksheet
        workbook = writer.book
        worksheet = writer.sheets['Jobs Template']

        # Style the header row
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        # Apply header styling
        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment

        # Adjust column widths
        column_widths = {
            'A': 25,  # Customer
            'B': 20,  # Customer Reference No
            'C': 35,  # Department/Person In Charge/Sub-Customer
            'D': 20,  # Service
            'E': 15,  # Vehicle
            'F': 20,  # Driver
            'G': 20,  # Contractor
            'H': 20,  # Vehicle Type
            'I': 15,  # Pickup Date
            'J': 12,  # Pickup Time
            'K': 30,  # Pickup Location
            'L': 30,  # Drop-off Location
            'M': 20,  # Passenger Name
            'N': 18,  # Passenger Mobile
            'O': 40   # Remarks
        }

        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        # Add data validation for dropdowns (max 255 characters for Excel formula)
        # Customer dropdown (Column A)
        if validation_data['customers']:
            customer_formula = f'"{",".join(validation_data["customers"][:50])}"'  # Limit to 50 to avoid Excel formula length limit
            customer_validation = DataValidation(type="list", formula1=customer_formula, allow_blank=True)
            customer_validation.add('A2:A1000')
            worksheet.add_data_validation(customer_validation)

        # Service dropdown (Column D)
        if validation_data['services']:
            service_formula = f'"{",".join(validation_data["services"][:50])}"'
            service_validation = DataValidation(type="list", formula1=service_formula, allow_blank=True)
            service_validation.add('D2:D1000')
            worksheet.add_data_validation(service_validation)

        # Vehicle dropdown (Column E)
        if validation_data['vehicles']:
            vehicle_formula = f'"{",".join(validation_data["vehicles"][:50])}"'
            vehicle_validation = DataValidation(type="list", formula1=vehicle_formula, allow_blank=True)
            vehicle_validation.add('E2:E1000')
            worksheet.add_data_validation(vehicle_validation)

        # Driver dropdown (Column F)
        if validation_data['drivers']:
            driver_formula = f'"{",".join(validation_data["drivers"][:50])}"'
            driver_validation = DataValidation(type="list", formula1=driver_formula, allow_blank=True)
            driver_validation.add('F2:F1000')
            worksheet.add_data_validation(driver_validation)

        # Contractor dropdown (Column G)
        if validation_data['contractors'] and validation_data['contractors'][0]:
            contractor_formula = f'"{",".join(validation_data["contractors"][:50])}"'
            contractor_validation = DataValidation(type="list", formula1=contractor_formula, allow_blank=True)
            contractor_validation.add('G2:G1000')
            worksheet.add_data_validation(contractor_validation)

        # Vehicle Type dropdown (Column H)
        if validation_data['vehicle_types'] and validation_data['vehicle_types'][0]:
            vehicle_type_formula = f'"{",".join(validation_data["vehicle_types"][:50])}"'
            vehicle_type_validation = DataValidation(type="list", formula1=vehicle_type_formula, allow_blank=True)
            vehicle_type_validation.add('H2:H1000')
            worksheet.add_data_validation(vehicle_type_validation)

        # Freeze header row
        worksheet.freeze_panes = 'A2'

    print(f"\nExcel file created successfully: {output_filename}")
    print(f"Total jobs: {len(df)}")
    print(f"Valid jobs: 90")
    print(f"Invalid jobs (for testing): 10")


def main():
    """Main function to generate test data"""
    print("=" * 80)
    print("BULK UPLOAD TEST DATA GENERATOR")
    print("=" * 80)
    print("\nGenerating 100 test jobs for bulk upload functionality testing...")
    print("(90 valid jobs + 10 jobs with validation errors)")
    print()

    try:
        # Generate test data
        df, validation_data = generate_test_jobs(100)

        # Create Excel file
        output_file = 'bulk_upload_100_jobs_test.xlsx'
        create_excel_with_formatting(df, validation_data, output_file)

        print("\n" + "=" * 80)
        print("SUCCESS!")
        print("=" * 80)
        print(f"\nFile generated: {os.path.abspath(output_file)}")
        print("\nYou can now use this file to test the bulk upload functionality.")
        print("The file contains:")
        print("  - 90 valid job entries with diverse test data")
        print("  - 10 entries with intentional validation errors")
        print("  - Dropdown data validation for key fields")
        print("  - Proper Excel formatting and styling")
        print("\n" + "=" * 80)

    except Exception as e:
        print(f"\nERROR: {str(e)}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == '__main__':
    main()
